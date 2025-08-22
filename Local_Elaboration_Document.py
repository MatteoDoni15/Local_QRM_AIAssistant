# Standard library  
import os  
import io  
from io import BytesIO  
import time    
import re  
import base64  
import json  
import logging  
import uuid  
from datetime import datetime, timezone  
from typing import List, Any, Dict, Optional, Awaitable, Callable   
from dotenv import load_dotenv


# Third-party libraries  
import requests  
import fitz  
import numpy as np  
import tiktoken  
import aiohttp  
import asyncio  
import grapheme  
from PIL import Image  
import pandas as pd 
  
# Azure SDK  
#import azure.functions as func  
from azure.core.credentials import AzureKeyCredential  
#from azure.keyvault.secrets import SecretClient  
#from azure.identity import DefaultAzureCredential  
#from azure.ai.textanalytics import TextAnalyticsClient 
from azure.ai.textanalytics.aio import TextAnalyticsClient #importandolo da .aio è asincrono il metodo
#from azure.ai.vision.imageanalysis import ImageAnalysisClient
from azure.ai.vision.imageanalysis.aio import ImageAnalysisClient 
from azure.ai.vision.imageanalysis.models import VisualFeatures  
from azure.core.exceptions import HttpResponseError, ServiceRequestError, ServiceResponseError    

# OpenAI / Mistral SDK  
import openai
#from openai import AzureOpenAI  
from openai import AsyncAzureOpenAI  
#from mistralai import Mistral  
#from mistralai.utils.retries import RetryConfig, BackoffStrategy
#from mistralai import  OCRPageObject



import yake
import langid  

##########LIBRERIE DA AGGIUNGERE
from langchain.text_splitter import MarkdownHeaderTextSplitter
from langchain_text_splitters import RecursiveCharacterTextSplitter
import markdown2 
import gc


# Docling imports
from docling.document_converter import DocumentConverter
from docling.datamodel.base_models import InputFormat
from docling.datamodel.pipeline_options import PdfPipelineOptions
from docling.document_converter import PdfFormatOption

# Tesseract OCR
import pytesseract

# Docling imports
from docling.document_converter import DocumentConverter
from docling.datamodel.base_models import InputFormat
from docling.datamodel.pipeline_options import (
    PdfPipelineOptions, 
    PipelineOptions,
    TesseractCliOcrOptions, 
    EasyOcrOptions
)
from docling.document_converter import PdfFormatOption
from docling.datamodel.base_models import InputFormat
from docling.datamodel.pipeline_options import PdfPipelineOptions
from docling.document_converter import DocumentConverter, PdfFormatOption
from docling.utils.export import generate_multimodal_pages
from docling.utils.utils import create_hash
from docling.datamodel.document import DocumentStream
from docling.chunking import HybridChunker
from docling_core.transforms.chunker.tokenizer.openai import OpenAITokenizer


from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

import openpyxl
from docling.backend.pypdfium2_backend import PyPdfiumDocumentBackend
from docling.datamodel.base_models import InputFormat
from docling.document_converter import (
    DocumentConverter,
    PdfFormatOption,
    WordFormatOption,)
from docling.pipeline.simple_pipeline import SimplePipeline
from docling.pipeline.standard_pdf_pipeline import StandardPdfPipeline
from docling_core.types import DoclingDocument
from docling.document_converter import ConversionResult


#### configuration ocr-tesseract
import sys
import shutil



load_dotenv()

EMBEDDING_DIM= 1536



#############################################################################################################################################################################################################
#############################################################################################################################################################################################################
#############################################################################################################################################################################################################
#############################################################################################################################################################################################################
#############################################################################################################################################################################################################







AZURE_STORAGE_CONNECTION_STRING = os.getenv("AZURE_STORAGE_CONNECTION_STRING")
AZURE_BLOB_CONTAINER = os.getenv("AZURE_BLOB_CONTAINER")


AZURE_SEARCH_ENDPOINT = os.getenv("AZURE_SEARCH_ENDPOINT")
AZURE_NAME_SEARCH_SERVICE= os.getenv("AZURE_NAME_SEARCH_SERVICE")
AZURE_SEARCH_KEY = os.getenv("AZURE_SEARCH_KEY")



####NUOVO INDICIZZATORE
AZURE_INDEXER_NAME= os.getenv("AZURE_INDEXER_NAME")


#AZURE_SEARCH_INDEX = "test_qrmaiprojectindex_2"

AZURE_SEARCH_INDEX= os.getenv("AZURE_SEARCH_INDEX")
 

 
AZURE_NAME_SEARCH_SERVICE= os.getenv("AZURE_NAME_SEARCH_SERVICE")
AZURE_SEARCH_KEY =   os.getenv("AZURE_SEARCH_KEY")
 
 
AZURE_OPENAI_EMBEDDING_DEPLOYMENT= os.getenv("AZURE_OPENAI_EMBEDDING_DEPLOYMENT")
AZURE_OPENAI_API_KEY=os.getenv("AZURE_OPENAI_API_KEY")
AZURE_OPENAI_ENDPOINT =os.getenv("AZURE_OPENAI_ENDPOINT")
AZURE_OPENAI_LLM_DEPLOYMENT =os.getenv("AZURE_OPENAI_LLM_DEPLOYMENT")
AZURE_OPENAI_API_VERSION =os.getenv("AZURE_OPENAI_API_VERSION")
 
 
MISTRAL_API_KEY=os.getenv("MISTRAL_API_KEY") 
DEPLOYMENT_MISTRAL_OCR = os.getenv("DEPLOYMENT_MISTRAL_OCR")
MISTRAL_ENDPOINT = os.getenv("MISTRAL_ENDPOINT")
 
 
 
API_KEY_COGNITIVE_SERVICE=os.getenv("API_KEY_COGNITIVE_SERVICE")
ENDPOINT_COGNITIVE_SERVICE= os.getenv("ENDPOINT_COGNITIVE_SERVICE")





DATA_SOURCE_NAME =os.getenv("DATA_SOURCE_NAME") 


CONTAINER_NAME = os.getenv("CONTAINER_NAME")

####EmbeddingModel è un client diverso dall'openAI dell'llm perchè è stato fatto in un'altra regione
AZURE_OPENAI_EMBEDDING_MODEL=os.getenv("AZURE_OPENAI_EMBEDDING_MODEL")
AZURE_OPENAI_EMBEDDING_DEPLOYMENT= os.getenv("AZURE_OPENAI_EMBEDDING_DEPLOYMENT")
ENDPOINT_EMBEDDING= os.getenv("ENDPOINT_EMBEDDING")
API_KEY_EMBEDDING= os.getenv("API_KEY_EMBEDDING")
API_VERSION_EMBEDDING = os.getenv("API_VERSION_EMBEDDING")

####Endpoint LLM
AZURE_OPENAI_API_KEY= os.getenv("AZURE_OPENAI_API_KEY")
AZURE_OPENAI_ENDPOINT = os.getenv("AZURE_OPENAI_ENDPOINT")
AZURE_OPENAI_LLM_DEPLOYMENT = os.getenv("AZURE_OPENAI_LLM_DEPLOYMENT")
AZURE_OPENAI_LLM_API_VERSION=os.getenv("AZURE_OPENAI_LLM_API_VERSION")


API_ENDPOINT_PROMPT_FLOW = os.getenv("API_ENDPOINT_PROMPT_FLOW")
API_KEY_PROMPT_FLOW = os.getenv("API_KEY_PROMPT_FLOW")


####MistralEndpoint
MISTRAL_API_KEY=os.getenv("MISTRAL_API_KEY")
DEPLOYMENT_MISTRAL_OCR =os.getenv("DEPLOYMENT_MISTRAL_OCR")
MISTRAL_ENDPOINT = os.getenv("MISTRAL_ENDPOINT")



API_KEY_COGNITIVE_SERVICE= os.getenv("API_KEY_COGNITIVE_SERVICE")
ENDPOINT_COGNITIVE_SERVICE= os.getenv("ENDPOINT_COGNITIVE_SERVICE")




AZURE_FUNCTION_URL= os.getenv("AZURE_FUNCTION_URL")
 
AZURE_FUNCTION_KEY= os.getenv("AZURE_FUNCTION_KEY")


AZURE_FUNCTION_URL_LOCAL=os.getenv("AZURE_FUNCTION_URL_LOCAL")
AZURE_FUNCTION_KEY_LOCAL= os.getenv("AZURE_FUNCTION_KEY_LOCAL")






#############################################################################################################################################################################################################
#############################################################################################################################################################################################################
#############################################################################################################################################################################################################
#############################################################################################################################################################################################################
#############################################################################################################################################################################################################




EMBEDDING_DIM= 1536


GLOBAL_TIMEOUT = 900  

GLOBAL_YAKE_EXTRACTORS_CACHE = {}

IMAGE_RESOLUTION_SCALE = 2.0

# Define common YAKE! parameters
YAKE_COMMON_PARAMS = {
    "n": 3,
    "top": 10,
    "dedupLim": 0.95,
    "dedupFunc": 'seqm'
}
logging.info("YAKE! Keyword Extractor parameters defined. Instances will be cached dynamically per language.")



# Configurazione Tesseract (Docker/ENV/Windows) con auto-rilevamento
# Replaces previous simple Docker-only configuration

def _configure_tesseract()-> bool:


    cmd_env = os.getenv("TESSERACT_CMD")
    tessdata_env = os.getenv("TESSDATA_PREFIX")

    # 1) ENV override / 2) PATH / 3) fallback
    found_path = None
    if cmd_env and os.path.exists(cmd_env):
        found_path = cmd_env
        logging.info(f"Tesseract from ENV: {cmd_env}")
    else:
        which_name = "tesseract.exe" if os.name == "nt" else "tesseract"
        which_path = shutil.which(which_name)
        if which_path:
            found_path = which_path
            logging.info(f"Tesseract found on PATH: {which_path}")
        else:
            candidates = []
            if sys.platform.startswith("win"):
                localapp = os.environ.get("LOCALAPPDATA", "")
                candidates = [
                    r"C:\Program Files\Tesseract-OCR\tesseract.exe",
                    r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe",
                ]
                if localapp:
                    candidates.append(os.path.join(localapp, "Programs", "Tesseract-OCR", "tesseract.exe"))
            elif sys.platform.startswith("darwin"):
                candidates = [
                    "/opt/homebrew/bin/tesseract",
                    "/usr/local/bin/tesseract",
                    "/opt/local/bin/tesseract",
                ]
            else:
                candidates = [
                    "/usr/bin/tesseract",
                    "/usr/local/bin/tesseract",
                ]
            for p in candidates:
                if os.path.exists(p):
                    found_path = p
                    logging.info(f"Tesseract found at {p}")
                    break

    if found_path:
        pytesseract.pytesseract.tesseract_cmd = found_path
        return True
    else:
        logging.warning("Tesseract executable not found. Set TESSERACT_CMD or add tesseract to PATH.")
        return False





#################################################################################################################################################################
#################################################################################################################################################################




#################################################################################################################################################################
#################################################################################################################################################################



def check_parent_ids(record_ids: List[str], chat_id:str):  
    """  
    Checks the existence of given parent IDs in the Azure Cognitive Search index.  
  
    Args:  
        record_ids (list of str): A list of parent ID strings to check in the Azure Search index.  
        chat_id (str): The chat_id to filter by.  
    Returns:  
        tuple:  
            found (list of str): Parent IDs found in the Azure Search index.  
            not_found (list of str): Parent IDs not found in the Azure Search index.  
    """  
  
    found = []  
    not_found = []  
    headers = {  
        "Content-Type": "application/json",  
        "api-key": AZURE_SEARCH_KEY  
    }  
    url = f"https://{AZURE_NAME_SEARCH_SERVICE}.search.windows.net/indexes/{AZURE_SEARCH_INDEX}/docs/search?api-version=2024-07-01"  
    # Query che cerca solo questi parent_id e lo chat_id specificato  
    ids_odata = " or ".join([f"parent_id eq '{pid}'" for pid in record_ids])     
    # Combinazione dei filtri con l'operatore AND  
    filter_query = f"chat_id eq '{chat_id}' and ({ids_odata})"
    logging.info(f"check_parent_ids: {filter_query} ")   
    body = {  
        "search": "*",  
        "filter": filter_query,  
        "select": "parent_id"  
    }  
    response = requests.post(url, headers=headers, json=body)  
    
    results = response.json()  
    logging.info("Risposta:\n%s", json.dumps(results, indent=2, ensure_ascii=False))  
    returned_ids = set(doc["parent_id"] for doc in results.get("value", []))  
    for pid in record_ids:  
        if pid in returned_ids:  
            found.append(pid) 
            logging.info(f"RecordId {pid} trovato") 
        else:  
            not_found.append(pid)
            logging.info(f"RecordId {pid} non trovato")  
    return found, not_found 



async def wait_for_record_ids_async(results,  record_ids: list,log_suffix:str,global_timeout_time: float, timeout_seconds=215, poll_interval=5   ):  
    started = time.time()  
    not_found = list(set(record_ids)) 
    processed_records = set()  # Track which records have been processed

    while not_found and (time.time() - started < timeout_seconds):  
        if time.time() >= global_timeout_time:
            logging.warning(f"{log_suffix} Global timeout reached, stopping processing")
            break
        
        try:
            chat_id = results[0].get("chat_id", "") if results else ""  
            if chat_id == "":
                logging.error(f"{log_suffix} No chat_id found in results, cannot proceed with record ID checks")
                break      
            found, not_found_now = check_parent_ids(record_ids=not_found, chat_id=chat_id) 
            logging.info(f"Found: {found} and not found {not_found_now}")
            
            # Track newly found records
            newly_found = set(found) - processed_records
            processed_records.update(found)
            
            # Update only newly found records to avoid race conditions
            for item in results:
                if item["RecordId"] in newly_found:
                    logging.info(f"Marking as processed RecordId: {item['RecordId']}")  
                    item["is_processed"] = True
                    
            if found:  
                logging.info(f"Found {len(found)} RecordId: {found}")  
            if not_found_now:  
                logging.info(f"Still to find {len(not_found_now)} RecordId: {not_found_now}")  
            if not_found_now:  
                await asyncio.sleep(poll_interval)  
            not_found = not_found_now
        except Exception as e:
            logging.error(f"{log_suffix} Error in wait_for_record_ids_async: {e}")
            break
            
    return not_found, results   





#################################################################################################################################################################
#################################################################################################################################################################
################################################################################################################################################################
################################################################################################################################################################

async def describe_image_with_gpt41_async(base64_image: str,
                                           client: AsyncAzureOpenAI,
                                             sem: asyncio.Semaphore,
                                              global_timeout_time: float,
                                               timeout: int = 120) -> str:
    """  
    Asynchronously generates a brief image description using GPT-4.1 via Azure OpenAI.  
  
    Args:  
        base64_image (str): Base64-encoded PNG image string.  
        client (AsyncAzureOpenAI): Async Azure OpenAI client.  
        semaphore (asyncio.Semaphore): Semaphore to limit concurrent requests.  
  
    Returns:  
        str: Short description of the image or an error message.  
    """  
    prefix_log= f"describe_image_with_gpt41_async: "

    system = """You are a virtual assistant who extracts information from images of PDF pages using OCR.  
    Your goal is to convert the recognized content into a structured Markdown format.  

    Your answer must always include the following three sections, strictly in this order:

    ### Markdown Text  
    This section contains the plain text content extracted from the page, formatted in Markdown. If there are images, insert them in the appropriate place using the syntax:
    `![]({image_id})`  
    `<!-- {description} -->`  
    where `{image_id}` must follow the format `image_{page_idx + 1}_{image_counter}`. The image ID is generated programmatically and refers to the order of appearance on the current page.

    ### Page Description  
    Provide a short description of the layout and contents of the page, including any notable visual elements such as charts, diagrams, or decorative images.

    ### Markdown Tables  
    Convert any table found on the page into valid Markdown table format.  
    If there are no tables found, write: `No tables found.`

    Be accurate and preserve the visual and semantic structure of the original page as much as possible. """
    message = [{"role": "system", "content": system}]
    message.append({
        "role": "user", "content": [
            {"type": "text", "text": "Analyze the following pdf image:\n"},
            {
                "type": "image_url",
                "image_url": {"url": f"data:image/jpeg;base64,{base64_image}"},
            }
        ]
    })

    if global_timeout_time:  
        time_left = global_timeout_time - time.time()  
        if time_left <= 0:  
            logging.error(f"{prefix_log}Global timeout expired.")  
            return ""  
    else:  
        time_left = timeout  # fallback  

    effective_timeout = min(timeout, time_left) 


    try:
        async with sem: 
            response = await asyncio.wait_for(
                client.chat.completions.create(
                    model=AZURE_OPENAI_LLM_DEPLOYMENT,
                    messages=message
                ),
                timeout=effective_timeout
            )
            return response.choices[0].message.content
    except openai.RateLimitError as e:  
        logging.error(f"{prefix_log}Rate limit reached (429): {e}")  
        return ""  
  
    except openai.APITimeoutError as e:  
        logging.error(f"{prefix_log}Request timeout: {e}")  
        return ""  
  
    except openai.APIConnectionError as e:  
        logging.error(f"{prefix_log}API connection error: {e}")  
        return ""  
  
    except openai.AuthenticationError as e:  
        logging.error(f"{prefix_log}Authentication error: {e}")  
        return ""  
  
    except openai.PermissionDeniedError as e:  
        logging.error(f"{prefix_log}Permission denied: {e}")  
        return ""  
  
    except openai.BadRequestError as e:  
        logging.error(f"{prefix_log}Bad request (400): {e}")  
        return ""  
  
    except openai.NotFoundError as e:  
        logging.error(f"{prefix_log}Resource not found (404): {e}")  
        return ""  
  
    except openai.UnprocessableEntityError as e:  
        logging.error(f"{prefix_log}Unprocessable entity (422): {e}")  
        return ""
    except asyncio.CancelledError:  
        logging.warning(f"{prefix_log}Task was cancelled")  
        return ""


async def get_info_image_with_azure_async(client_ImageAnalysis:ImageAnalysisClient ,
                                           image_stream,
                                             sem_ocr: asyncio.Semaphore,
                                             global_timeout_time: float,
                                             timeout: int = 120
                                             ) -> str:
    """
    Analyzes an image stream asynchronously using Azure Image Analysis to extract various visual features.

    This function utilizes Azure Cognitive Services to perform optical character recognition (OCR), 
    tagging, object detection, and people detection on the provided image. It respects a semaphore 
    to control the concurrency of requests to the Azure service.

    Args:
        client_ImageAnalysis (ImageAnalysisClient): An authenticated Azure Image Analysis client.
        image_stream: The image data as a stream (e.g., bytes, file-like object).
        sem_ocr (asyncio.Semaphore): An asyncio semaphore to limit concurrent OCR operations.

    Returns:
        str: A formatted string containing the extracted information. 
             This includes recognized text, tags, detected objects, and the count of people.
             Returns "Image content not clearly identifiable" if no significant information is found.

    Notes:
        - The function currently uses `VisualFeatures.READ`, `VisualFeatures.TAGS`, 
          `VisualFeatures.OBJECTS`, and `VisualFeatures.PEOPLE`.
        - Features like `CAPTION` and `DENSE_CAPTIONS` are commented out due to regional limitations 
          (e.g., 'Sweden Central') but are included as a placeholder for future use.
        - Only tags and objects with a confidence score greater than 0.6 are included in the output.
        - For people detection, it reports both the count of high-confidence detections and total detections.
    """

    if global_timeout_time:  
        time_left = global_timeout_time - time.time()  
        if time_left <= 0:  
            logging.error("get_info_image_with_azure_async: Global timeout expired.")  
            return ""  
    else:  
        time_left = timeout  # fallback  
  
    effective_timeout = min(timeout, time_left)  


    try:
        async with sem_ocr: 
            result = await asyncio.wait_for( 
                client_ImageAnalysis.analyze(
                image_data=image_stream,
                visual_features=[
                    #VisualFeatures.CAPTION, ### purtroppo non supportato nella regione sweden central ma apppena lo rendono utilizzabile da Usare
                    #VisualFeatures.DENSE_CAPTIONS,  ### purtroppo non supportato nella regione sweden central ma apppena lo rendono utilizzabile da Usare
                    VisualFeatures.READ,
                    VisualFeatures.TAGS,
                    VisualFeatures.OBJECTS,
                    #VisualFeatures.SMART_CROPS, ###forse per cose future
                    VisualFeatures.PEOPLE
                            ]
                        ), 
                    timeout=effective_timeout
                    )

        output = []

    #    if result.caption and result.caption.confidence > 0.5:
    #        output.append(f"Caption: '{result.caption.text}'\n") ###, Confidence: {result.caption.confidence:.4f}

    #    if result.dense_captions and result.dense_caption.confidence > 0.5:
    #        output.append("Dense Captions:")
    #        for caption in result.dense_captions.list:
    #            output.append(f" Caption text: '{caption.text}'") ###, {caption.bounding_box}, Confidence: {caption.confidence:.4f}
    #        output.append("")

        if result.read:
            text_parts =[]
            output.append("Read:")
            for block in result.read.blocks:
                for line in block.lines:
                    if line.text.strip():  # Evita testo vuoto
                        text_parts.append(line.text.strip())
            if text_parts:
                output.append(f"Text: {' '.join(text_parts)}")


        if result.tags:
            high_conf_tags = [tag.name for tag in result.tags.list if tag.confidence > 0.6]
            if high_conf_tags:
                output.append(f"Tags: {', '.join(high_conf_tags)}") 

        if result.objects:
            objects = []
            for obj in result.objects.list:
                if obj.tags and obj.tags[0].confidence > 0.6:
                    objects.append(obj.tags[0].name)
            if objects:
                output.append(f"Objects: {', '.join(set(objects))}") 

        if result.people:
            people_count = len(result.people.list)
            high_conf_people = [p for p in result.people.list if p.confidence > 0.6]
            if high_conf_people:
                output.append(f"People: {len(high_conf_people)} detected and Total People detected {people_count}")

    #    if result.smart_crops:
    #        output.append("Smart Cropping:")
    #        for crop in result.smart_crops.list:
    #            output.append(f"  Aspect ratio: {crop.aspect_ratio}, Smart crop: {crop.bounding_box}")

        return "\n".join(output) if output else ""
    except asyncio.TimeoutError:  
        logging.error("get_info_image_with_azure_async: Azure ImageAnalysis request timed out.")  
        return ""  
    except asyncio.CancelledError:  
        logging.warning("get_info_image_with_azure_async: Task was cancelled")  
        return ""
    except Exception as e:  
        logging.error(f"get_info_image_with_azure_async: Errore durante l'analisi immagine Azure: {type(e).__name__}: {e}")  
        return "" 













#################################################################################################################################################################
#################################################################################################################################################################
#################################################################################################################################################################
#################################################################################################################################################################



def chunk_text_to_docs_extract_key_phrases_inputs(text, language, max_elements=5120, overlap=1000, batch_size=10):  
    """  
    Splits text into character-based chunks for Azure Text Analytics key phrase extraction,  
    wrapping each chunk as a dictionary ready for API input, with batching.  
  
    Args:  
        text (str): The input text to split into chunks.  
        language (str): The language code of the text.  
        max_elements (int): Maximum number of characters per chunk.  
        overlap (int): Number of overlapping characters between consecutive chunks.  
        batch_size (int): Number of items per batch sent to the API.  
  
    Returns:  
        list[list[dict]]: Batches of chunk dicts, each with an "id", "text", and "language".  
    """    
    docs = []  
    text_elements = list(grapheme.graphemes(text))  
    start = 0  
    while start < len(text_elements):  
        end = min(start + max_elements, len(text_elements))  
        chunk = ''.join(text_elements[start:end])  
        doc = {  
            "id": str(uuid.uuid4()),  
            "text": chunk,  
            "language": language  
        }  
        docs.append(doc)  
        start += max_elements - overlap  
  
    # Split docs in batch sottoliste di max batch_size elementi  
    docs_batches = [docs[i:i+batch_size] for i in range(0, len(docs), batch_size)]  
    return docs_batches



def detected_language_with_langid_async(text: str) -> str:  
    """ Detects language using langid as a fallback mechanism."""
    # Fallback sync con langid (non ha api async) 
    logging.info("Falback to langid for language detection")  
    try:  
        lang, score = langid.classify(text)  
        logging.info(f"detect_language_with_langid_async: Detected language via langid: {lang} (score: {score})")  
        return lang  
    except Exception as e:  
        logging.error(f"detect_language_with_langid_async: langid fallback also failed: {e}")  
        return "en"  




async def detect_language_with_fallback_async(text:str,
                                               text_analytics_client:TextAnalyticsClient,
                                               global_timeout_time: float,
                                               timeout: int = 120) -> str:  
    """Detects language using Azure Text Analytics (async), falls back to langid if Azure fails."""  

    if not text or not text.strip():
        logging.warning("Empty or whitespace-only text provided for language detection")
        return "en"

    if global_timeout_time:  
        time_left = global_timeout_time - time.time()  
        if time_left <= 0:  
            logging.error("detect_language_with_fallback_async: Global timeout expired.")  
            return "en"  # Default fallback language 
    else:  
        time_left = timeout  # fallback 

    effective_timeout = min(timeout, time_left) 

    if text_analytics_client:  
        try:  

            language_result = await asyncio.wait_for(text_analytics_client.detect_language(  
                documents=[{"id": "1", "text": text[:4900]}]  
                ), 
            timeout=effective_timeout 
            )              
            if (language_result 
                and len(language_result) > 0 
                and not language_result[0].is_error 
                and hasattr(language_result[0], "primary_language")
                and language_result[0].primary_language
                and hasattr(language_result[0].primary_language, "iso6391_name")):                              
              
                detected_language = language_result[0].primary_language.iso6391_name  
                confidence = getattr(language_result[0].primary_language, "confidence_score", 0)
                logging.info(f"Detected language via Azure: {detected_language} (confidence: {confidence:.2f})")   
                return detected_language  
            else:  
                logging.warning("detect_language_with_fallback_async: Azure failed to detect language or returned an error.") 
            return detected_language_with_langid_async(text)                 
        except Exception as e:  
            logging.error(f"detect_language_with_fallback_async: Azure language detection failed due to Exception: {e}")  
            return detected_language_with_langid_async(text)
        except asyncio.CancelledError:  
            logging.warning("detect_language_with_fallback_async: Task was cancelled")  
            return detected_language_with_langid_async(text)







# Utility function to extract keywords using YAKE! as a fallback
def extract_keywords_with_yake(text: str, language: str, num_keywords: int = 10) -> list[str]:
    """
    Extracts keywords from a text using YAKE! as a fallback mechanism.
    It uses a cache to store YAKE! KeywordExtractor instances per language,
    avoiding repeated initialization.
    """
    try:
        # Check if an extractor for this language already exists in the cache
        if language not in GLOBAL_YAKE_EXTRACTORS_CACHE:
            logging.info(f"Initializing YAKE! KeywordExtractor for language '{language}' and adding to cache.")
            # If not, create a new instance and store it
            GLOBAL_YAKE_EXTRACTORS_CACHE[language] = yake.KeywordExtractor(
                lan=language,
                n=YAKE_COMMON_PARAMS["n"],
                top=YAKE_COMMON_PARAMS["top"],
                dedupLim=YAKE_COMMON_PARAMS["dedupLim"],
                dedupFunc=YAKE_COMMON_PARAMS["dedupFunc"]
            )
        
        # Retrieve the extractor from the cache
        kw_extractor = GLOBAL_YAKE_EXTRACTORS_CACHE[language]
        
        keywords_with_scores = kw_extractor.extract_keywords(text)
        # YAKE! returns a list of (keyword, score) tuples. Extract only the keyword strings.
        # Note: In YAKE!, a lower score indicates higher relevance.
        return [kw for kw, score in keywords_with_scores[:num_keywords]]
    except Exception as e:
        # Log any errors during YAKE! execution
        logging.error(f"Error during keyword extraction with YAKE! (fallback) for language '{language}': {e}")
        return []  


async def extract_keywords_with_fallback_async(
    text: str,
    language: str,
    text_analytics_client: TextAnalyticsClient,
    global_timeout_time: float,
    timeout: int = 120,
    max_retries: int = 3,
    retry_wait_sec: int = 5
) -> list:
    """
    Extracts keywords using Azure Text Analytics with aggressive 429 handling.
    On rate limit (429), immediately falls back to YAKE without retries.
    
    Args:
        text: Text to analyze
        language: Language code (e.g., 'en', 'it')
        text_analytics_client: Azure Text Analytics client (can be None)
        max_retries: Maximum retry attempts for transient errors (not 429)
        retry_wait_sec: Wait time between retries
    
    Returns:
        list: List of unique keywords/key phrases
    """
    
    # Input validation
    if not text or not text.strip():
        logging.warning("extract_keywords_with_fallback_async: Empty or whitespace-only text provided for keyword extraction")
        return []




    # Try Azure Text Analytics first
    if text_analytics_client:
        list_documents_for_ta = chunk_text_to_docs_extract_key_phrases_inputs(
            text=text,
            language=language
        )

        # FIX: gestione corretta del time budget
        if global_timeout_time:
            time_left = global_timeout_time - time.time()
            if time_left <= 0:
                logging.error("extract_keywords_with_fallback_async: Global timeout expired.")
                return []
        else:
            time_left = timeout
        effective_timeout = min(timeout, time_left)

        all_keywords = []
        rate_limit_hit = False
        # Process each document batch
        for documents_batch in list_documents_for_ta:
            if rate_limit_hit:
                break
            retries = 0
            batch_success = False
            while retries < max_retries and not rate_limit_hit:
                if global_timeout_time and time.time() >= global_timeout_time:
                    logging.error("extract_keywords_with_fallback_async: Global timeout expired during batch processing")
                    return []
                try:
                    result = await asyncio.wait_for(
                        text_analytics_client.extract_key_phrases(documents=documents_batch),
                        timeout=effective_timeout
                    )
                    batch_keywords = []
                    for doc_result in result:
                        if not doc_result.is_error:
                            batch_keywords.extend(doc_result.key_phrases)
                        else:
                            logging.warning(f"extract_keywords_with_fallback_async: Document processing error: {doc_result.error}")
                    all_keywords.extend(batch_keywords)
                    batch_success = True
                    break
                except (HttpResponseError, ServiceRequestError, ServiceResponseError) as e:
                    if hasattr(e, "status_code") and e.status_code == 429:
                        logging.warning("extract_keywords_with_fallback_async: Azure rate limit (429). Falling back to YAKE.")
                        rate_limit_hit = True
                        break
                    msg = str(e).lower()
                    if ("timeout" in msg or "timed out" in msg or isinstance(e, ServiceRequestError)):
                        retries += 1
                        if retries < max_retries:
                            logging.warning(f"Timeout (attempt {retries}/{max_retries}): {e}")
                            await asyncio.sleep(retry_wait_sec)
                        else:
                            logging.error(f"Max retries reached for batch. Error: {e}")
                            break
                    else:
                        logging.error(f"Non-retryable Azure error: {e}")
                        break
                except Exception as e:
                    logging.error(f"Unexpected error in Azure extraction: {e}")
                    break
                except asyncio.CancelledError:
                    logging.warning("extract_keywords_with_fallback_async: Task was cancelled")
                    break
        if all_keywords and not rate_limit_hit:
            logging.info(f"Azure extraction successful. Found {len(all_keywords)} keywords.")
            return list(set(all_keywords))
        if rate_limit_hit:
            logging.info("Falling back to YAKE due to rate limit")
        else:
            logging.info("Azure returned no keywords, falling back to YAKE")
    else:
        logging.info("Text Analytics client not available, using YAKE fallback")

    if global_timeout_time and time.time() >= global_timeout_time:
        logging.error("Global timeout expired in extract_keywords_with_fallback_async")
        return []

    # Fallback to YAKE
    try:
        keywords = extract_keywords_with_yake(text, language)
        if keywords:
            logging.info(f"YAKE fallback successful. Found {len(keywords)} keywords.")
            return list(set(keywords))
        else:
            logging.warning("YAKE fallback did not find keywords")
    except Exception as e:
        logging.error(f"YAKE fallback failed: {e}")
    return []

async def process_single_chunk_for_keywords_async(chunk: Dict[str, Any],
                                                   detected_language: str,
                                                     text_analytics_client:TextAnalyticsClient,
                                                     global_timeout_time :float) -> Dict[str, Any]:  
    """
    Processes a single text chunk to extract keywords using an asynchronous Azure Text Analytics pipeline.

    This function takes a dictionary representing a text chunk, extracts its raw text content,
    and then attempts to identify key phrases using the `extract_keywords_with_fallback_async`
    function. It also incorporates any headers found within the chunk as keywords.

    Args:
        chunk (Dict[str, Any]): A dictionary containing the chunk's data. Expected keys include:
            - "text_raw" (str): The raw text content of the chunk.
            - "metadata" (Dict[str, Any]): Metadata associated with the chunk,
              potentially including a "page" number.
            - "base64_imgs_list" (List[str]): A list of Base64 encoded images found in the chunk.
            - "headers" (Dict[str, str]): A dictionary of headers found in the chunk,
              where keys are header levels (e.g., "h1", "h2") and values are header texts.
            - "text_markdown" (str): The markdown formatted text content of the chunk.
        detected_language (str): The language of the text content within the chunk (e.g., "en", "it").
        text_analytics_client (TextAnalyticsClient): An authenticated Azure Text Analytics client
                                                     (assumed to be an aio.TextAnalyticsClient instance).

    Returns:
        Dict[str, Any]: A dictionary containing the original chunk data augmented with:
            - "keywords" (List[str]): A list of unique keywords extracted from the text content
              and headers.
            - "detected_language" (str): The language that was used for keyword extraction.

    Notes:
        - Keyword extraction is skipped if `text_content` is empty or only contains whitespace.
        - The `extract_keywords_with_fallback_async` function handles the actual
          Azure Text Analytics API call and any fallback logic (e.g., to YAKE!).
        - Headers are directly added to the list of keywords.
        - The returned "keywords" list contains only unique entries.
    """
    text_content = chunk.get("text_raw", "")  
    page_number = chunk.get("metadata", {}).get("page", "unknown")  
    base64_imgs_list = chunk.get("base64_imgs_list", [])  
    headers= chunk.get("headers", {})
    logging.debug(f"Processing chunk page {page_number}, text length: {len(text_content)}")  
  
    key_phrases = []  

    if text_content and text_content.strip():   
        try:  
            key_phrases = await extract_keywords_with_fallback_async(  
                text= text_content,  
                language=detected_language,  
                text_analytics_client= text_analytics_client, 
                global_timeout_time=global_timeout_time 
            )  
        except Exception as e:  
            logging.error(f"Error during async key phrase extraction for page {page_number}: {e}")  
    else:  
        logging.info(f"Skipping keyword extraction for page {page_number}: empty text.") 

    if headers: 
        for header_level, header_text in chunk.get("headers").items():
            key_phrases.append(header_text)
    


    return {  
        "text_raw": text_content,
        "text_markdown": chunk.get("text_markdown", ""),
        "metadata": chunk.get("metadata", {}),  
        "base64_imgs_list": base64_imgs_list,  
        "keywords": list(set(key_phrases)),
        "detected_language": detected_language
    }  



def make_thumbnail_base64(base64_str, thumb_size=(200, 200), format="JPEG"):  


    logging.info(base64_str[:20])
    # Rimuovi eventuale header base64  
    if base64_str.startswith('data:image'):  
        base64_str = base64_str.split(',', 1)[1]  
    logging.info(base64_str[:20])

    try:  
        image_bytes = base64.b64decode(base64_str)  
    except Exception as e:  
        logging.error(f"Error in decodification base64: {e}")  
        raise  
  
    try:  
        with Image.open(io.BytesIO(image_bytes)) as img:  
            orig_size = img.size  
            logging.info(f"Original image size (pixels): {orig_size}")  
            logging.info(f"Original image size (bytes): {len(image_bytes)}")  
  
            img = img.convert("RGB")  
            img.thumbnail(thumb_size)  
            thumb_size_actual = img.size  
            logging.info(f"Thumbnail size (pixels): {thumb_size_actual}")  
  
            buf = io.BytesIO()  
            img.save(buf, format=format)  
            thumb_bytes = buf.getvalue()  
            logging.info(f"Thumbnail size (bytes): {len(thumb_bytes)}")  
  
            # Calcola la riduzione percentuale  
            perc_pixel = 100 * (thumb_size_actual[0] * thumb_size_actual[1]) / (orig_size[0] * orig_size[1])  
            perc_bytes = 100 * len(thumb_bytes) / len(image_bytes)  
            logging.info(f"Riduction pixel: {perc_pixel:.1f}% from the original size")  
            logging.info(f"Riduction bytes: {perc_bytes:.1f}% from the original size")  
  
            thumb_base64 = base64.b64encode(thumb_bytes).decode('utf-8')  
            return thumb_base64  
  
    except Exception as e:  
        logging.error(f"Error in opening or converting image: {e}")  
        with open("immagine_non_valida.bin", "wb") as f:  
            f.write(image_bytes)  
        raise   

def _detect_docling_input_format(fn: str) -> Optional[InputFormat]:
    # Se fn è un URL http/https, ritorna HTML
    if isinstance(fn, str) and fn.strip().lower().startswith(("http://", "https://")):
        fmt = getattr(InputFormat, "HTML", None)
        try:
            logging.info(f"_detect_docling_input_format: filename='{fn}', rilevato URL, formato_rilevato='HTML'")
            print(f"_detect_docling_input_format: filename='{fn}', rilevato URL, formato_rilevato='HTML'")
        except Exception:
            pass
        return fmt

    ext_ = (os.path.splitext(fn or "")[1] or "").lower().lstrip(".")
    mapping = {
        "pdf": "PDF", "docx": "DOCX","xlsm": "XLSX", "xlsx": "XLSX", "pptx": "PPTX",
        "md": "MD", "adoc": "ASCIIDOC", "asciidoc": "ASCIIDOC",
        "html": "HTML", "xhtml": "HTML", "csv": "CSV",
        "png": "IMAGE", "jpg": "IMAGE", "jpeg": "IMAGE",
        "tif": "IMAGE", "tiff": "IMAGE", "bmp": "IMAGE", "webp": "IMAGE",
    }
    key = mapping.get(ext_)
    fmt = getattr(InputFormat, key, None) if key else None
    try:
        logging.info(f"_detect_docling_input_format: filename='{fn}', estensione='{ext_}', formato_rilevato='{fmt}'")
        print(f"_detect_docling_input_format: filename='{fn}', estensione='{ext_}', formato_rilevato='{fmt}'")
    except Exception:
        pass
    return fmt


def get_chunk_docling_document(conv_result: ConversionResult, input_fmt: InputFormat=InputFormat.ASCIIDOC):
    # Encoder GPT‑4.1 (o200k_base)
    enc = tiktoken.get_encoding("o200k_base")
    openai_tok = OpenAITokenizer(
        tokenizer=enc,
        max_tokens=128 * 1024,  # 128k contesto
    )
    if input_fmt == InputFormat.XLSX:
        max_tokens= 2000
        overlap_tokens= 100
    elif input_fmt == InputFormat.CSV:
        max_tokens= 1500
        overlap_tokens= 50
    elif input_fmt == InputFormat.PPTX:
        max_tokens= 1200
        overlap_tokens= 100
    elif input_fmt == InputFormat.PDF:
        max_tokens= 800
        overlap_tokens= 150
    else:
        max_tokens= 600
        overlap_tokens= 100


    # Passa il tokenizer al chunker; fallback a callable se l’API richiede token_counter
    try:
        chunker = HybridChunker(
            max_tokens=max_tokens,
            overlap_tokens=overlap_tokens,
            tokenizer=openai_tok,
        )
    except TypeError:
        chunker = HybridChunker(
            max_tokens=max_tokens,
            overlap_tokens=overlap_tokens,
            token_counter=lambda s: len(enc.encode(s)),
        )

    return chunker.chunk(conv_result.document)    





def _get_page_withdocling_pdf(document_io: io.BytesIO, tesseract_config: bool= True, document_name:str = "Unknown"):
    prefix_logging = f"_get_page_withdocling_pdf: filename='{document_name}'"

    logging.info(f"{prefix_logging} - starting")
    input_doc_path = DocumentStream(stream=document_io, name=document_name)
    logging.info(f"{prefix_logging} - created DocumentStream.")
    if tesseract_config:
        pipeline_options = PdfPipelineOptions()
        pipeline_options.images_scale = IMAGE_RESOLUTION_SCALE
        pipeline_options.generate_page_images = True

        ocr_options = TesseractCliOcrOptions(force_full_page_ocr=True, lang=["fra","deu","ita", "eng", "spa","por"])
        pipeline_options.ocr_options = ocr_options
        logging.info(f"{prefix_logging} - PipelineOptions e tesseract-OCR configuration.")
    else:
        pipeline_options = PdfPipelineOptions()
        pipeline_options.do_ocr = True
        pipeline_options.do_table_structure = True
        pipeline_options.table_structure_options.do_cell_matching = True
        pipeline_options.images_scale = IMAGE_RESOLUTION_SCALE
        pipeline_options.generate_page_images = True
        logging.info(f"{prefix_logging} - PipelineOptions e EASYOCR configuration.")

   
    doc_converter = DocumentConverter(allowed_formats=[InputFormat.PDF],
        format_options={
            InputFormat.PDF: PdfFormatOption(pipeline_options=pipeline_options)
        }
    )
    logging.info(f"{prefix_logging} - setted DocumentConverter.")

    logging.info(f"{prefix_logging} - Starting conversion.")
    conv_res = doc_converter.convert(input_doc_path)
    logging.info(f"{prefix_logging} - Document conversion completed.")

    rows = []
    for (
        content_text,
        content_md,
        content_dt,
        page_cells,
        page_segments,
        page,
    ) in generate_multimodal_pages(conv_res):
        dpi = page._default_image_scale * 72
        logging.info(f"{prefix_logging} - Elaborazione pagina {page.page_no+1} (dpi={dpi}).")
        rows.append(
            {
                "document": conv_res.input.file.name,
                "hash": conv_res.input.document_hash,
                "page_hash": create_hash(
                    conv_res.input.document_hash + ":" + str(page.page_no - 1)
                ),
                "image": {
                    "width": page.image.width,
                    "height": page.image.height,
                    "bytes": page.image.tobytes(),
                },
                "cells": page_cells,
                "contents": content_text,
                "raw_markdown": content_md,
                "contents_dt": content_dt,
                "segments": page_segments,
                "page": page.page_no + 1,
                "extra": {
                    "page_num": page.page_no + 1,
                    "width_in_points": page.size.width,
                    "height_in_points": page.size.height,
                    "dpi": dpi,
                },
                "conv_result": conv_res,           
                "input_format": InputFormat.PDF,

            }
        )

    logging.info(f"{prefix_logging} - Document converted and multimodal pages.")

    return rows


def _get_page_with_selenium(url: str):
    chrome_options = Options()
    chrome_options.add_argument("--headless")  
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("--no-sandbox")

    # Usa webdriver-manager per scaricare il driver giusto
    service = Service(ChromeDriverManager().install())

    driver = webdriver.Chrome(service=service, options=chrome_options)

    try:
        driver.get(url)

        # esempio: attesa di un elemento (da cambiare con xpath reale)
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.TAG_NAME, "body"))
        )

        return driver.page_source
    finally:
        driver.quit()


def _convert_html_with_docling(html: str, filename: str):
    # trasformo la stringa HTML in uno stream di byte
    stream = BytesIO(html.encode("utf-8"))

    # creo un DocumentStream
    ds = DocumentStream(stream=stream, name=filename)

    # creo il converter
    converter = DocumentConverter(allowed_formats=[InputFormat.HTML])

    # converto
    result = converter.convert(ds)
    

    return result

def _get_page_withdocling_excel(document_io: io.BytesIO, document_name: str = "Unknown"):
    prefix_logging = f"_get_page_withdocling_excel: filename='{document_name}'"
    logging.info(f"{prefix_logging} - starting")
    
    # Carica workbook in memoria
    wb = openpyxl.load_workbook(document_io, data_only=True)
    logging.info(f"{prefix_logging} - loaded workbook with sheets: {wb.sheetnames}")
    
    # Inizializza Docling
    doc_converter = DocumentConverter(
        allowed_formats=[InputFormat.XLSX]
    )

    logging.info(f"{prefix_logging} - DocumentConverter initialized")
    
    rows = []
    for idx, sheet_name in enumerate(wb.sheetnames, start=1):
        sheet = wb[sheet_name]
        # Crea BytesIO temporaneo con solo questo sheet
        temp_wb_io = io.BytesIO()
        temp_wb = openpyxl.Workbook()
        temp_sheet = temp_wb.active
        temp_sheet.title = sheet_name
        
        for i, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            for j, value in enumerate(row, start=1):
                temp_sheet.cell(row=i, column=j, value=value)
        
        temp_wb.save(temp_wb_io)
        temp_wb_io.seek(0)
        
        # Converte sheet con Docling
        input_doc_stream = DocumentStream(stream=temp_wb_io, name=f"{document_name}_{sheet_name}.xlsx")
        conv_res = doc_converter.convert(input_doc_stream)
        # Genera “pagina” dallo sheet (sheet = pagina)
        logging.info(f"{prefix_logging} - processed sheet '{sheet_name}' as page 1")

        md = conv_res.document.export_to_markdown() if hasattr(conv_res.document, "export_to_markdown") else None
        txt = conv_res.document.text if hasattr(conv_res.document, "text") else None
        #### c'è la ripetizione contents_md e raw_markdown
        rows.append({
            "document": conv_res.input.file.name,
            "hash": conv_res.input.document_hash,
#            "page_hash": create_hash(conv_res.input.document_hash + ":" + str(0)),
            "image": [],  # non abbiamo immagini dagli Excel
            "cells": conv_res.document.cells if hasattr(conv_res.document, "cells") else None,
            "contents": txt,
#            "contents_md": md,
            "contents_dt": conv_res.document.export_to_dict() if hasattr(conv_res.document, "export_to_dict") else None,
            "segments": None,
            "page": idx,                 # <-- intero
            "sheet_name": sheet_name,    # <-- conserva il nome del foglio
            "raw_markdown": md or (txt or ""),  # <-- chiave uniforme usata a valle
            "conv_result": conv_res,           # <-- aggiunto
            "input_format": InputFormat.XLSX,
        })

    logging.info(f"{prefix_logging} - Excel converted and sheets processed as pages")    
    return rows 
    





def _get_withdocling_custom(document_io: io.BytesIO, document_name: str = "Unknown")-> ConversionResult:

    input_doc_path = DocumentStream(stream=document_io, name=document_name)

    doc_converter = (
        DocumentConverter(  # all of the below is optional, has internal defaults.
            allowed_formats=[
                InputFormat.PDF,
                InputFormat.IMAGE,
                InputFormat.DOCX,
                InputFormat.HTML,
                InputFormat.PPTX,
                InputFormat.ASCIIDOC,
                InputFormat.CSV,
                InputFormat.MD,
            ],  # whitelist formats, non-matching files are ignored.
            format_options={
                InputFormat.PDF: PdfFormatOption(
                    pipeline_cls=StandardPdfPipeline, backend=PyPdfiumDocumentBackend
                ),
                InputFormat.DOCX: WordFormatOption(
                    pipeline_cls=SimplePipeline  # , backend=MsWordDocumentBackend
                ),
            },
        )
    )

    conv_results = doc_converter.convert(input_doc_path)
    return conv_results



async def docling_convert_to_pages_async(
    file_obj: io.BytesIO,
    filename: str,
    detected_language: str,
    tesseract_config: bool,
    global_timeout_time: float = None,
) -> List[Dict[str, Any]]:

    prefix_log= "docling_convert_to_pages_async: "


    if global_timeout_time and time.time() >= global_timeout_time - 5:
        logging.warning(f"{prefix_log} Timeout budget too small")
        print(f"{prefix_log} Timeout budget too small")
        return [{"page": 1, "raw_markdown": "", "input_format": None, "conv_result": None}]


    #ext = (os.path.splitext(filename or "")[1] or "").lower()
    working_file_obj = file_obj
    working_filename = filename



    input_fmt = _detect_docling_input_format(working_filename)
    if not input_fmt:
        logging.error(f"{prefix_log} Unsupported format {working_filename}")
        print(f"{prefix_log} Unsupported format {working_filename}")
        return [{"page": 1, "raw_markdown": "", "input_format": None, "conv_result": None}]

    pre_lang = None
    try:
        if input_fmt == InputFormat.PDF:
            try:
                pdf_bytes = working_file_obj.getvalue()
                with fitz.open(stream=pdf_bytes, filetype="pdf") as doc_ps:
                    sample_pages = []
                    max_pages = min(3, len(doc_ps))
                    for i in range(max_pages):
                        try:
                            txt = doc_ps[i].get_text("text") or ""
                        except Exception:
                            txt = ""
                        if txt:
                            sample_pages.append(txt.strip())
                        if sum(len(s) for s in sample_pages) > 4000:
                            break
                    sample_text = " ".join(sample_pages)[:5000]
                if sample_text.strip():
                    lang_code, score = langid.classify(sample_text)
                    if score >= 0.80:  # soglia semplice
                        pre_lang = lang_code[:2]
                        if pre_lang not in ["fra","deu","ita", "eng", "spa","por"]:
                            tesseract_config= False
                        logging.info(f"{prefix_log} pre-sampling PDF language='{pre_lang}' (score={score:.2f})")
                        print(f"{prefix_log} pre-sampling PDF language='{pre_lang}' (score={score:.2f})")
                    else:
                        logging.info(f"{prefix_log} pre-sampling PDF low confidence (score={score:.2f}) keep='{detected_language}'")
                        print(f"{prefix_log} pre-sampling PDF low confidence (score={score:.2f}) keep='{detected_language}'")
            except Exception as e_ps_pdf:
                logging.warning(f"{prefix_log} PDF presample failed: {e_ps_pdf}")
                print(f"{prefix_log} PDF presample failed: {e_ps_pdf}")

            return _get_page_withdocling_pdf(document_io=  file_obj, tesseract_config= tesseract_config, document_name = filename)
            
        elif input_fmt == InputFormat.HTML and isinstance(filename, str) and filename.strip().lower().startswith(("http://", "https://")):
            html = _get_page_with_selenium(filename)
            result = _convert_html_with_docling(html, filename)
            return [{"page": 1, "raw_markdown": result.document.export_to_markdown(), "input_format": InputFormat.HTML, "conv_result":result}]



        elif input_fmt == InputFormat.XLSX:
            return  _get_page_withdocling_excel(document_io=file_obj, document_name=filename)
            
        else:
            doc=  _get_withdocling_custom(document_io= file_obj, document_name= filename)
            return [{"page": 1, "raw_markdown": doc.document.export_to_markdown(), "input_format": input_fmt, "conv_result": doc}]

    except Exception as e:
        logging.error(f"{prefix_log} Error occurred: {e}")
        print(f"{prefix_log} Error occurred: {e}")
        return [{"page": 1, "raw_markdown": "", "input_format": None, "conv_result": None}]



async def _extract_images_and_chunk_single_page(
    page_number: int,
    page_markdown: str,
    client_azure_openai: AsyncAzureOpenAI,
    client_ImageAnalysis: ImageAnalysisClient,
    sem: asyncio.Semaphore,
    sem_ocr: asyncio.Semaphore,
    global_timeout_time: float,
    input_format: InputFormat,
    conv_result: ConversionResult,
    sheet_name: str= "",
) -> List[Dict[str, Any]]:
    """
    Estrae immagini inline base64 da una singola pagina markdown,
    genera thumbnail, crea task per descrizioni (GPT-4.1 se pagina senza testo, Azure Image Analysis se disponibile),
    sostituisce con placeholder + descrizione, poi usa split_markdown_and_raw e
    restituisce una lista di chunk con metadata.page = page_number.
    """
    if input_format == InputFormat.PDF or input_format == InputFormat.DOCX or input_format == InputFormat.HTML or input_format == InputFormat.MD:
        prefix_log = f"_extract_images_and_chunk_single_page:"
        logging.info(f"{prefix_log} start page={page_number} length_markdown={len(page_markdown)}")
        print(f"{prefix_log} start page={page_number} length_markdown={len(page_markdown)}")

        text = page_markdown
        page_idx = page_number - 1  # solo per formattazione id coerente con snippet

        # Rileva immagini inline base64
        pattern = re.compile(
            r"!\[(?P<alt>.*?)\]\((?P<src>data:image/(?:png|jpeg|jpg|gif|webp|tiff|bmp);base64,[A-Za-z0-9+/=\s]+)\)",
            re.IGNORECASE
        )

        image_description_tasks: List[tuple[str, asyncio.Future]] = []
        matches_to_process: List[tuple[int, int, Optional[str]]] = []
        taskid_to_base64: Dict[str, str] = {}
        image_id_to_base64: Dict[str, str] = {}

        # Prepara task per descrizioni immagini
        for match in pattern.finditer(text):
            data_uri = match.group("src")
            raw_b64 = data_uri.split(",", 1)[1] if "," in data_uri else ""
            if not raw_b64:
                matches_to_process.append((match.start(), match.end(), None))
                continue

            try:
                base64_thumb = make_thumbnail_base64(raw_b64, thumb_size=(200, 200), format="JPEG")
            except Exception:
                # fallback: usa l’originale
                base64_thumb = raw_b64

            task_id = f"image_task_{page_idx}_{len(image_description_tasks)}"
            taskid_to_base64[task_id] = base64_thumb

            # Se la pagina è “vuota” → GPT-4.1 Vision
            if text.strip() == "":
                image_description_tasks.append(
                    (task_id, describe_image_with_gpt41_async(
                        base64_image=base64_thumb,
                        client=client_azure_openai,
                        sem=sem,
                        global_timeout_time=global_timeout_time
                    ))
                )
            # Altrimenti prova Azure Image Analysis (se disponibile)
            elif client_ImageAnalysis:
                try:
                    image_bytes = base64.b64decode(base64_thumb)
                    try:
                        img = Image.open(io.BytesIO(image_bytes))
                        img.verify()
                    except Exception as e:
                        raise ValueError("The base64 data is not a decodable image: " + str(e))
                    image_stream = io.BytesIO(image_bytes)
                    image_description_tasks.append(
                        (task_id, get_info_image_with_azure_async(
                            client_ImageAnalysis=client_ImageAnalysis,
                            image_stream=image_stream.getvalue(),
                            sem_ocr=sem_ocr,
                            global_timeout_time=global_timeout_time
                        ))
                    )
                except Exception as e:
                    logging.error(f"{prefix_log} Error preparing image stream for Azure Image Analysis: {e}")
                    image_description_tasks.append((task_id, asyncio.sleep(0, result=f"[Error in image preparation: {e}]")))
            else:
                # Fallback: se Azure non disponibile, usa GPT-4.1 Vision
                image_description_tasks.append(
                    (task_id, describe_image_with_gpt41_async(
                        base64_image=base64_thumb,
                        client=client_azure_openai,
                        sem=sem,
                        global_timeout_time=global_timeout_time
                    ))
                )

            matches_to_process.append((match.start(), match.end(), task_id))

        # Time budget
        time_left = global_timeout_time - time.time() if global_timeout_time else 120
        if time_left <= 5:
            logging.warning(f"{prefix_log} Global timeout less than 5 seconds, stopping processing")
            return []

        # Esegui i task in parallelo
        image_descriptions: Dict[str, str] = {}
        if image_description_tasks:
            try:
                results = await asyncio.wait_for(
                    asyncio.gather(*[task for _, task in image_description_tasks]),
                    timeout=time_left
                )
            except asyncio.TimeoutError:
                logging.warning(f"{prefix_log} Global timeout reached while waiting image_description_tasks")
                return []
            except asyncio.CancelledError:
                logging.warning(f"{prefix_log} Task was cancelled!")
                return []
            except Exception as e:
                logging.error(f"{prefix_log} Error during image_description_tasks: {e}")
                return []

            for (tid, _), result in zip(image_description_tasks, results):
                image_descriptions[tid] = result or ""

        # Ricostruisci il testo con descrizioni inserite
        updated_text_parts = []
        last_end = 0
        image_count_on_page = 0

        for start, end, task_id in matches_to_process:
            updated_text_parts.append(text[last_end:start])

            if task_id and task_id in image_descriptions:
                image_count_on_page += 1
                description = image_descriptions[task_id]
                image_id = f"image_{page_idx + 1}_{image_count_on_page}"
                base64_thumb = taskid_to_base64.get(task_id)
                if base64_thumb:
                    image_id_to_base64[image_id] = base64_thumb

                if not description or description.strip() == "":
                    updated_text_parts.append(
                        f"![]({image_id})\n<!-- This is a general description and you need to understand the image from the text above and below -->\n"
                    )
                else:
                    updated_text_parts.append(
                        f"![]({image_id})\n<!-- {description.strip()} -->\n"
                    )
            else:
                # niente descrizione per questa immagine → testo originale
                updated_text_parts.append(text[start:end])

            last_end = end

        updated_text_parts.append(text[last_end:])
        updated_text = "".join(updated_text_parts)

        # Split in chunk e collega le immagini presenti in ogni chunk
        image_pattern = re.compile(r'!\[\]\((image_\d+_\d+)\)')
        results: List[Dict[str, Any]] = []

        for small_chunk in split_markdown_and_raw(updated_text):
            md_chunk = small_chunk.get("markdown", "")
            img_ids_in_chunk: List[str] = []
            for m in image_pattern.finditer(md_chunk):
                img_id = m.group(1)
                if img_id not in img_ids_in_chunk:
                    img_ids_in_chunk.append(img_id)
            base64_imgs_list = [image_id_to_base64[i] for i in img_ids_in_chunk if i in image_id_to_base64]

            results.append({
                "text_raw": small_chunk.get("raw", ""),
                "text_markdown": md_chunk,
                "metadata": {"page": page_number, "sheet_name": sheet_name},
                "base64_imgs_list": base64_imgs_list,
                "headers": small_chunk.get("headers", {}),
            })

        logging.info(f"{prefix_log} page={page_number} immagini_rilevate={len(taskid_to_base64)} chunk_creati={len(results)}")
        print(f"{prefix_log} page={page_number} immagini_rilevate={len(taskid_to_base64)} chunk_creati={len(results)}")
        return results
    elif input_format == InputFormat.XLSX:
        prefix_log = f"_extract_images_and_chunk_single_page[XLSX]:"
        logging.info(f"{prefix_log} start sheet='{sheet_name}' page={page_number} length_markdown={len(page_markdown) if page_markdown else 0}")
        print(f"{prefix_log} start sheet='{sheet_name}' page={page_number} length_markdown={len(page_markdown) if page_markdown else 0}")

        text = page_markdown or ""
        page_idx = page_number - 1

        # Rileva immagini inline base64 nel markdown (se presenti)
        pattern = re.compile(
            r"!\[(?P<alt>.*?)\]\((?P<src>data:image/(?:png|jpeg|jpg|gif|webp|tiff|bmp);base64,[A-Za-z0-9+/=\s]+)\)",
            re.IGNORECASE
        )

        image_description_tasks: List[tuple[str, asyncio.Future]] = []
        matches_to_process: List[tuple[int, int, Optional[str]]] = []
        taskid_to_base64: Dict[str, str] = {}
        image_id_to_base64: Dict[str, str] = {}

        for match in pattern.finditer(text):
            data_uri = match.group("src")
            raw_b64 = data_uri.split(",", 1)[1] if "," in data_uri else ""
            if not raw_b64:
                matches_to_process.append((match.start(), match.end(), None))
                continue

            try:
                base64_thumb = make_thumbnail_base64(raw_b64, thumb_size=(200, 200), format="JPEG")
            except Exception:
                base64_thumb = raw_b64

            task_id = f"image_task_{page_idx}_{len(image_description_tasks)}"
            taskid_to_base64[task_id] = base64_thumb

            if client_ImageAnalysis:
                try:
                    image_bytes = base64.b64decode(base64_thumb)
                    try:
                        img = Image.open(io.BytesIO(image_bytes))
                        img.verify()
                    except Exception as e:
                        raise ValueError("The base64 data is not a decodable image: " + str(e))
                    image_stream = io.BytesIO(image_bytes)
                    image_description_tasks.append(
                        (task_id, get_info_image_with_azure_async(
                            client_ImageAnalysis=client_ImageAnalysis,
                            image_stream=image_stream.getvalue(),
                            sem_ocr=sem_ocr,
                            global_timeout_time=global_timeout_time
                        ))
                    )
                except Exception as e:
                    logging.error(f"{prefix_log} Error preparing image stream for Azure Image Analysis: {e}")
                    image_description_tasks.append((task_id, asyncio.sleep(0, result=f"[Error in image preparation: {e}]")))
            else:
                image_description_tasks.append(
                    (task_id, describe_image_with_gpt41_async(
                        base64_image=base64_thumb,
                        client=client_azure_openai,
                        sem=sem,
                        global_timeout_time=global_timeout_time
                    ))
                )

            matches_to_process.append((match.start(), match.end(), task_id))

        # Time budget
        time_left = global_timeout_time - time.time() if global_timeout_time else 120
        if time_left <= 5:
            logging.warning(f"{prefix_log} Global timeout less than 5 seconds, stopping processing")
            return []

        # Esegui i task in parallelo
        image_descriptions: Dict[str, str] = {}
        if image_description_tasks:
            try:
                results = await asyncio.wait_for(
                    asyncio.gather(*[task for _, task in image_description_tasks]),
                    timeout=time_left
                )
            except asyncio.TimeoutError:
                logging.warning(f"{prefix_log} Global timeout reached while waiting image_description_tasks")
                return []
            except asyncio.CancelledError:
                logging.warning(f"{prefix_log} Task was cancelled!")
                return []
            except Exception as e:
                logging.error(f"{prefix_log} Error during image_description_tasks: {e}")
                return []

            for (tid, _), result in zip(image_description_tasks, results):
                image_descriptions[tid] = result or ""

        # Ricostruisci markdown con placeholder e descrizioni
        updated_text_parts = []
        last_end = 0
        image_count_on_page = 0

        for start, end, task_id in matches_to_process:
            updated_text_parts.append(text[last_end:start])

            if task_id and task_id in image_descriptions:
                image_count_on_page += 1
                description = image_descriptions[task_id]
                image_id = f"image_{page_idx + 1}_{image_count_on_page}"
                base64_thumb = taskid_to_base64.get(task_id)
                if base64_thumb:
                    image_id_to_base64[image_id] = base64_thumb

                if not description or description.strip() == "":
                    updated_text_parts.append(
                        f"![]({image_id})\n<!-- This is a general description and you need to understand the image from the text above and below -->\n"
                    )
                else:
                    updated_text_parts.append(
                        f"![]({image_id})\n<!-- {description.strip()} -->\n"
                    )
            else:
                updated_text_parts.append(text[start:end])

            last_end = end

        updated_text_parts.append(text[last_end:])
        updated_text = "".join(updated_text_parts)

        # Chunk Excel-safe (preserva tabelle)
        md_chunks = chunk_markdown_excel_safe(
            updated_text,
            max_tokens=2000,
            overlap_tokens=100,
            enc_name="o200k_base"
        )

        image_pattern = re.compile(r'!\[\]\((image_\d+_\d+)\)')
        results: List[Dict[str, Any]] = []

        for md_chunk in md_chunks:
            # Collega immagini presenti nel chunk
            img_ids_in_chunk: List[str] = []
            for m in image_pattern.finditer(md_chunk):
                img_id = m.group(1)
                if img_id not in img_ids_in_chunk:
                    img_ids_in_chunk.append(img_id)
            base64_imgs_list = [image_id_to_base64[i] for i in img_ids_in_chunk if i in image_id_to_base64]

            results.append({
                "text_raw": markdown_to_raw(md_chunk),
                "text_markdown": md_chunk,
                "metadata": {"page": page_number, "sheet_name": sheet_name},
                "base64_imgs_list": base64_imgs_list,
                "headers": {},  # niente header parsing per XLSX
            })

        logging.info(f"{prefix_log} sheet='{sheet_name}' page={page_number} immagini_rilevate={len(image_id_to_base64)} chunk_creati={len(results)}")
        print(f"{prefix_log} sheet='{sheet_name}' page={page_number} immagini_rilevate={len(image_id_to_base64)} chunk_creati={len(results)}")
        return results
    elif input_format == InputFormat.CSV:
        prefix_log = f"_extract_images_and_chunk_single_page[CSV]:"
        logging.info(f"{prefix_log} start page={page_number} length_markdown={len(page_markdown) if page_markdown else 0}")
        print(f"{prefix_log} start page={page_number} length_markdown={len(page_markdown) if page_markdown else 0}")

        text = page_markdown or ""

        # Chunk table-safe come XLSX (CSV non ha immagini inline tipicamente)
        md_chunks = chunk_markdown_excel_safe(
            text,
            max_tokens=1500,
            overlap_tokens=50,
            enc_name="o200k_base"
        )
        results: List[Dict[str, Any]] = []
        for md_chunk in md_chunks:
            results.append({
                "text_raw": markdown_to_raw(md_chunk),
                "text_markdown": md_chunk,
                "metadata": {"page": page_number},
                "base64_imgs_list": [],
                "headers": {},
            })
        logging.info(f"{prefix_log} page={page_number} chunk_creati={len(results)}")
        print(f"{prefix_log} page={page_number} chunk_creati={len(results)}")
        return results
    
    elif input_format == InputFormat.PPTX:
        prefix_log = f"_extract_images_and_chunk_single_page[PPTX]:"
        logging.info(f"{prefix_log} start page={page_number} length_markdown={len(page_markdown) if page_markdown else 0}")
        print(f"{prefix_log} start page={page_number} length_markdown={len(page_markdown) if page_markdown else 0}")

        text = page_markdown or ""
        page_idx = page_number - 1

        pattern = re.compile(
            r"!\[(?P<alt>.*?)\]\((?P<src>data:image/(?:png|jpeg|jpg|gif|webp|tiff|bmp);base64,[A-Za-z0-9+/=\s]+)\)",
            re.IGNORECASE
        )

        image_description_tasks: List[tuple[str, asyncio.Future]] = []
        matches_to_process: List[tuple[int, int, Optional[str]]] = []
        taskid_to_base64: Dict[str, str] = {}
        image_id_to_base64: Dict[str, str] = {}

        for match in pattern.finditer(text):
            data_uri = match.group("src")
            raw_b64 = data_uri.split(",", 1)[1] if "," in data_uri else ""
            if not raw_b64:
                matches_to_process.append((match.start(), match.end(), None))
                continue

            try:
                base64_thumb = make_thumbnail_base64(raw_b64, thumb_size=(200, 200), format="JPEG")
            except Exception:
                base64_thumb = raw_b64

            task_id = f"image_task_{page_idx}_{len(image_description_tasks)}"
            taskid_to_base64[task_id] = base64_thumb

            if client_ImageAnalysis:
                try:
                    image_bytes = base64.b64decode(base64_thumb)
                    try:
                        img = Image.open(io.BytesIO(image_bytes)); img.verify()
                    except Exception as e:
                        raise ValueError("The base64 data is not a decodable image: " + str(e))
                    image_stream = io.BytesIO(image_bytes)
                    image_description_tasks.append(
                        (task_id, get_info_image_with_azure_async(
                            client_ImageAnalysis=client_ImageAnalysis,
                            image_stream=image_stream.getvalue(),
                            sem_ocr=sem_ocr,
                            global_timeout_time=global_timeout_time
                        ))
                    )
                except Exception as e:
                    logging.error(f"{prefix_log} Error preparing image stream for Azure Image Analysis: {e}")
                    image_description_tasks.append((task_id, asyncio.sleep(0, result=f"[Error in image preparation: {e}]")))
            else:
                image_description_tasks.append(
                    (task_id, describe_image_with_gpt41_async(
                        base64_image=base64_thumb,
                        client=client_azure_openai,
                        sem=sem,
                        global_timeout_time=global_timeout_time
                    ))
                )

            matches_to_process.append((match.start(), match.end(), task_id))

        time_left = global_timeout_time - time.time() if global_timeout_time else 120
        if time_left <= 5:
            logging.warning(f"{prefix_log} Global timeout less than 5 seconds, stopping processing")
            return []

        image_descriptions: Dict[str, str] = {}
        if image_description_tasks:
            try:
                results = await asyncio.wait_for(
                    asyncio.gather(*[task for _, task in image_description_tasks]),
                    timeout=time_left
                )
            except asyncio.TimeoutError:
                logging.warning(f"{prefix_log} Global timeout reached while waiting image_description_tasks")
                return []
            except asyncio.CancelledError:
                logging.warning(f"{prefix_log} Task was cancelled!")
                return []
            except Exception as e:
                logging.error(f"{prefix_log} Error during image_description_tasks: {e}")
                return []

            for (tid, _), result in zip(image_description_tasks, results):
                image_descriptions[tid] = result or ""

        updated_text_parts = []
        last_end = 0
        image_count_on_page = 0

        for start, end, task_id in matches_to_process:
            updated_text_parts.append(text[last_end:start])
            if task_id and task_id in image_descriptions:
                image_count_on_page += 1
                description = image_descriptions[task_id]
                image_id = f"image_{page_idx + 1}_{image_count_on_page}"
                base64_thumb = taskid_to_base64.get(task_id)
                if base64_thumb:
                    image_id_to_base64[image_id] = base64_thumb
                if not description or description.strip() == "":
                    updated_text_parts.append(f"![]({image_id})\n<!-- This is a general description and you need to understand the image from the text above and below -->\n")
                else:
                    updated_text_parts.append(f"![]({image_id})\n<!-- {description.strip()} -->\n")
            else:
                updated_text_parts.append(text[start:end])
            last_end = end

        updated_text_parts.append(text[last_end:])
        updated_text = "".join(updated_text_parts)

        # Chunk slide-aware per PPTX
        md_chunks = chunk_markdown_pptx_safe(
            updated_text,
            max_tokens=1200,
            overlap_tokens=100,
            enc_name="o200k_base"
        )

        image_pattern = re.compile(r'!\[\]\((image_\d+_\d+)\)')
        results: List[Dict[str, Any]] = []
        for md_chunk in md_chunks:
            img_ids_in_chunk: List[str] = []
            for m in image_pattern.finditer(md_chunk):
                img_id = m.group(1)
                if img_id not in img_ids_in_chunk:
                    img_ids_in_chunk.append(img_id)
            base64_imgs_list = [image_id_to_base64[i] for i in img_ids_in_chunk if i in image_id_to_base64]
            results.append({
                "text_raw": markdown_to_raw(md_chunk),
                "text_markdown": md_chunk,
                "metadata": {"page": page_number, "sheet_name": sheet_name},
                "base64_imgs_list": base64_imgs_list,
                "headers": {},
            })

        logging.info(f"{prefix_log} page={page_number} immagini_rilevate={len(image_id_to_base64)} chunk_creati={len(results)}")
        print(f"{prefix_log} page={page_number} immagini_rilevate={len(image_id_to_base64)} chunk_creati={len(results)}")
        return results

    else:
        temp_result = get_chunk_docling_document(conv_result=conv_result, input_fmt=input_format)
        for a in temp_result:
            results.append({
                "text_raw": a.text,
                "text_markdown": a.text,
                "metadata": {"page": page_number, "sheet_name": sheet_name},
                "base64_imgs_list":  [],
                "headers": {},
            })
        return results


#### 1
async def process_docling_pages_to_chunks_async(
    pages: List[Dict[str, Any]],
    client_azure_openai: AsyncAzureOpenAI = None,
    client_ImageAnalysis: ImageAnalysisClient = None,
    sem: asyncio.Semaphore = None,
    sem_ocr: asyncio.Semaphore = None,
    global_timeout_time: float = None
) -> List[Dict[str, Any]]:
    """
    Seconda fase: prende la lista di pagine (markdown) e produce i chunk finali per l'indicizzazione
    con estrazione immagini e chunking per pagina.
    """
    prefix_logging = "process_docling_pages_to_chunks_async: "

    logging.info(f"{prefix_logging} start num_pages={len(pages)}")
    print(f"{prefix_logging} start num_pages={len(pages)}")
    results: List[Dict[str, Any]] = []

    for p in pages:
        if global_timeout_time and time.time() >= global_timeout_time - 5:
            logging.warning(f"{prefix_logging} timeout budget nearly exhausted")
            print(f"{prefix_logging} timeout budget nearly exhausted")
            break
        pn = p["page"]
        md = p["raw_markdown"]
        conv_res= p["conv_result"]
        input_format= p["input_format"]
        try:
            page_chunks = await _extract_images_and_chunk_single_page(
                page_number=pn,
                page_markdown=md,
                client_azure_openai=client_azure_openai,
                client_ImageAnalysis=client_ImageAnalysis,
                sem=sem or asyncio.Semaphore(1),
                sem_ocr=sem_ocr or asyncio.Semaphore(1),
                global_timeout_time=global_timeout_time or (time.time() + 120),
                conv_result=conv_res,
                input_format=input_format,
                sheet_name=p.get("sheet_name", "")
            )
            results.extend(page_chunks)
        except Exception as e:
            logging.error(f"{prefix_logging} error on page {pn}: {e}")
            print(f"{prefix_logging} error on page {pn}: {e}")

    logging.info(f"{prefix_logging} totale_chunk={len(results)}")
    print(f"{prefix_logging} totale_chunk={len(results)}")
    return results


async def add_keywords_docling_tesseract_async(
    file_obj: io.BytesIO,
    filename: str,
    client_azure_openai: AsyncAzureOpenAI,
    text_analytics_client: TextAnalyticsClient,
    client_ImageAnalysis: ImageAnalysisClient,
    global_timeout_time: float,
    tesseract_config: bool,
    sem: asyncio.Semaphore,
    sem_ocr: asyncio.Semaphore,
) -> List[Dict[str, Any]]:
    """
    Nuova pipeline a due fasi (Docling pagine → elaborazione per pagina → keywords).
    """
    prefix_log= f"add_keywords_docling_tesseract_async: filename='{filename}'"

    if global_timeout_time and time.time() >= global_timeout_time - 5:
        logging.warning(f"{prefix_log} global timeout near")
        print(f"{prefix_log} global timeout near")
        return []

    # Fase 1: Docling -> pagine
    logging.info(f"{prefix_log} start")
    print(f"{prefix_log} start")
    pages_info = await docling_convert_to_pages_async(
        file_obj=file_obj,
        filename=filename,
        detected_language="en",  # lingua preferita per OCR (verrà rilevata poi)
        global_timeout_time=global_timeout_time,
        tesseract_config=tesseract_config,
    )
    if pages_info == [{"page": 1, "raw_markdown": "", "input_format": None, "conv_result": None}]:
        logging.error(f"{prefix_log} no pages returned by Docling")
        print(f"{prefix_log} no pages returned by Docling")
        return []

    # Fase 2: per pagina estrazione immagini + chunk
    if global_timeout_time and time.time() >= global_timeout_time - 5:
        logging.warning(f"{prefix_log} timeout before page processing")
        return []
    processed_chunks = await process_docling_pages_to_chunks_async(
        pages=pages_info,
        global_timeout_time=global_timeout_time,
        client_azure_openai=client_azure_openai,
        client_ImageAnalysis=client_ImageAnalysis,
        sem=sem,
        sem_ocr=sem_ocr
    )
    if not processed_chunks:
        logging.error(f"{prefix_log} no chunks after per-page processing")
        print(f"{prefix_log} no chunks after per-page processing")
        return []
    logging.info(f"{prefix_log} chunks_generati={len(processed_chunks)}")
    print(f"{prefix_log} chunks_generati={len(processed_chunks)}")

    # Rilevazione lingua su testo aggregato
    if global_timeout_time and time.time() >= global_timeout_time - 5:
        logging.warning(f"{prefix_log} timeout before language detection")
        return []
    all_text = " ".join([c.get("text_raw", "") for c in processed_chunks if c.get("text_raw")])
    detected_language = "en"
    if all_text.strip():
        try:
            detected_language = await detect_language_with_fallback_async(
                text=all_text[:4900],
                text_analytics_client=text_analytics_client,
                global_timeout_time=global_timeout_time
            )
            logging.info(f"{prefix_log} detected language {detected_language}")
            print(f"{prefix_log} detected language {detected_language}")
        except Exception as e:
            logging.error(f"{prefix_log} language detection error {e}")
            print(f"{prefix_log} language detection error {e}")
    else:
        logging.warning(f"{prefix_log} empty text for language detection")
        print(f"{prefix_log} empty text for language detection")

    # Estrazione keywords per chunk
    if global_timeout_time and time.time() >= global_timeout_time - 5:
        logging.warning(f"{prefix_log} timeout before keyword extraction")
        return []
    time_left = (global_timeout_time - time.time()) if global_timeout_time else 120
    if time_left <= 5:
        logging.warning(f"{prefix_log} insufficient time for keywords")
        return []

    try:
        keyword_tasks = [
            process_single_chunk_for_keywords_async(
                chunk=chunk,
                detected_language=detected_language,
                text_analytics_client=text_analytics_client,
                global_timeout_time=global_timeout_time
            )
            for chunk in processed_chunks
        ]
        final_docs = await asyncio.wait_for(asyncio.gather(*keyword_tasks), timeout=time_left)
        logging.info(f"{prefix_log} keyword_extraction completata documenti={len(final_docs)}")
        print(f"{prefix_log} keyword_extraction completata documenti={len(final_docs)}")
        if final_docs:
            logging.info(f"{prefix_log} primo_doc_keys={list(final_docs[0].keys())}")
            print(f"{prefix_log} primo_doc_keys={list(final_docs[0].keys())}")
        return final_docs
    except asyncio.TimeoutError:
        logging.warning(f"{prefix_log} timeout during keyword extraction")
        print(f"{prefix_log} timeout during keyword extraction")
        return []
    except Exception as e:
        logging.error(f"{prefix_log} keyword extraction error {e}")
        print(f"{prefix_log} keyword extraction error {e}")
        return []

def batch_by_payload_size(dicts, max_bytes=14*1024*1024):  # 14MB margine  
    """
    A generator that batches a list of dictionaries into smaller lists,
    ensuring that the JSON-encoded size of each batch does not exceed a specified maximum.

    This is useful for preparing payloads for APIs that have size limitations on requests.

    Args:
        dicts (list): A list of dictionaries to be batched.
        max_bytes (int, optional): The maximum allowed size (in bytes) for the JSON-encoded
                                   payload of a single batch. Defaults to 14MB.

    Yields:
        list: A list of dictionaries forming a batch, whose total JSON-encoded size
              is less than or equal to `max_bytes`.
    """
    batch = []  
    size = 0  
    for d in dicts:  
        s = len(json.dumps(d, ensure_ascii=False).encode('utf-8'))  
        if batch and size+s > max_bytes:  
            yield batch  
            batch = [d]  
            size = s  
        else:  
            batch.append(d)  
            size += s  
    if batch:  
        yield batch  

async def upload_batches(session, url, api_key, dicts_for_indexing):  
    """
    Asynchronously uploads batches of documents to a specified URL using HTTP POST requests.

    This function takes a list of dictionaries, batches them by payload size using `batch_by_payload_size`,
    and then sends each batch as a JSON payload to the target URL. It includes an API key in the headers
    for authentication. It logs the progress of the upload and raises an exception if any batch upload fails.

    Args:
        session (aiohttp.ClientSession): An aiohttp client session to make HTTP requests.
        url (str): The URL endpoint to which the batches will be uploaded.
        api_key (str): The API key to be included in the request headers for authentication.
        dicts_for_indexing (list): A list of dictionaries, where each dictionary represents a document
                                   to be indexed.

    Returns:
        list: A list of JSON responses received from the server for each successful batch upload.

    Raises:
        Exception: If a batch upload fails (i.e., the server responds with a status code other than 200).

    Logs:
        - INFO: For each batch uploaded, indicating the number of documents in the batch and the total uploaded count.
        - INFO: Once the entire upload process is completed.
    """
    headers = {'Content-Type': 'application/json', 'api-key': api_key}  
    uploaded = 0  
    responses = []  
    for batch in batch_by_payload_size(dicts_for_indexing):  
        payload = {"value": batch}  
        async with session.post(url, json=payload, headers=headers) as response:  
            if response.status != 200:  
                errtext = await response.text()  
                raise Exception(f"Batch upload failed: {response.status} {errtext}")  
            res = await response.json() 
            responses.append(res)  
            uploaded += len(batch)  
            logging.info(f"Uploaded a batch of {len(batch)} documents. Total uploaded: {uploaded}")  
    logging.info("Upload completed.") 
    return responses  







async def upload_index_docling_tesseract_from_url_async(
    blob_url: str,
    document_io: BytesIO,
    chat_id: str,
    parent_id: str,
    url: str, 
    embedding_function: Callable[[str, AsyncAzureOpenAI, str, asyncio.Semaphore, int, int, int], Awaitable[list[float]]],
    api_key: str, 
    client_embedding: AsyncAzureOpenAI,
    model_deployment: str,
    client: AsyncAzureOpenAI, 
    document_name: str,
    record_id: str,
    sem: asyncio.Semaphore,
    sem_download: asyncio.Semaphore,
    sem_ocr: asyncio.Semaphore,
    sem_embedding: asyncio.Semaphore,
    #client_mistral: Mistral,
    text_analytics_client: TextAnalyticsClient,
    client_ImageAnalysis: ImageAnalysisClient,
    global_timeout_time: float,
    tesseract_config: bool,
) -> bool:
    """
    Asynchronously downloads a document, processes it with Docling+Tesseract,
    generates embeddings, and uploads to Azure Search index.
    """
    prefix_log = f"upload_index_docling_tesseract_from_url_async: document_name='{document_name}'"
    #### aggiungo la possibilità di riscaricare il file se è nullo dall'url: non mi serve ma per generalità
    if document_name.strip().lower().startswith(("http://", "https://")):
        document_io= io.BytesIO()
    elif document_io is None or (hasattr(document_io, "getbuffer") and document_io.getbuffer().nbytes == 0) and not isinstance(document_name, str):
        # Download documento
        document_io = await get_file_with_retry(
            blob_url=blob_url, 
            sem_download=sem_download,
            RecordId=record_id,
            log_suffix="upload_index_docling_tesseract_from_url_async",
            global_timeout_time=global_timeout_time, 
            max_retries=3, 
            delay=3
        )
        if not document_io:
            logging.error(f"{prefix_log}: Failed to download document from {blob_url}. Aborting indexing.")
            return False
    else:
        try:
            document_io.seek(0)
        except Exception:
            pass

    if global_timeout_time and time.time() >= global_timeout_time - 5:  
        logging.error("Global timeout expired.")  
        return False

    try:
        # Processa documento con Docling+Tesseract
        processed_docs_with_keywords = await asyncio.wait_for(
            add_keywords_docling_tesseract_async(
                file_obj=document_io,
                filename=document_name,
                client_azure_openai=client, 
                text_analytics_client=text_analytics_client,
                client_ImageAnalysis=client_ImageAnalysis,
                global_timeout_time=global_timeout_time,
                tesseract_config= tesseract_config,
                sem=sem,
                sem_ocr=sem_ocr
            ),
            timeout=global_timeout_time - time.time()
        )
    
    except asyncio.TimeoutError:  
        logging.warning(f"{prefix_log}: Timeout in add_keywords_docling_tesseract_async. Skipping indexing.")  
        return False  
    except Exception as e:  
        logging.error(f"{prefix_log}: Error during add_keywords_docling_tesseract_async: {e}")  
        return False  

    if not processed_docs_with_keywords:
        logging.warning(f"{prefix_log}: No processed documents found for {document_name}. Skipping indexing.")
        return False

    # Resto del codice identico alla versione originale...
    dicts_for_indexing = []
    embedding_tasks = []
    docs_for_embedding = []

    if time.time() >= global_timeout_time:
        logging.warning(f"{prefix_log}: Global timeout reached, stopping processing")
        return False

    # Prepara embedding tasks
    for doc in processed_docs_with_keywords:
        page_number = str(doc.get("metadata", {}).get("page", "-1"))
        context_for_embedding = doc.get("text_raw", "")

        if not context_for_embedding or context_for_embedding.strip() == "":
            logging.warning(
                f"{prefix_log}: Skipping embedding for empty text_raw for document on page {page_number}."
            )
            doc["_skip_embedding_and_indexing"] = True 
            docs_for_embedding.append(doc)
            embedding_tasks.append(asyncio.sleep(0))
            continue

        docs_for_embedding.append(doc)
        embedding_tasks.append(
            embedding_function(
                text=context_for_embedding,
                global_timeout_time=global_timeout_time,
                client=client_embedding,
                model_deployment=model_deployment,
                sem_embedding=sem_embedding
            )
        )

    time_left = global_timeout_time - time.time()
    if time_left <= 5:
        logging.warning(f"{prefix_log}: Global timeout less than 5 seconds.")
        return False

    try:  
        embeddings = await asyncio.wait_for( 
            asyncio.gather(*embedding_tasks),
            timeout=time_left
        ) 
        if not embeddings or ["Blocking error"] in embeddings:
            logging.error(f"{prefix_log}: Blocking error detected. Skipping document from indexing.")
            return False

    except asyncio.TimeoutError:  
        logging.warning(f"{prefix_log}: Global timeout reached while waiting for embeddings.")  
        return False
    except Exception as e:  
        logging.error(f"{prefix_log}: Error during batch embedding: {e}")    
        return False

    if time.time() >= global_timeout_time:
        logging.warning(f"{prefix_log}: Global timeout reached, stopping processing")
        return False

    # Prepara documenti per l'indicizzazione
    for i, doc in enumerate(docs_for_embedding):
        if doc.get("_skip_embedding_and_indexing", False):
            continue
        if doc.get("metadata", {}).get("sheet_name", None):
            page_number = str(doc.get("metadata", {}).get("sheet_name"))
        else:
            page_number = str(doc.get("metadata", {}).get("page", "-1"))
        embedding = embeddings[i]

        if isinstance(embedding, Exception):
            logging.error(f"{prefix_log}: Error calculating embedding for page {page_number}: {embedding}. Skipping.")
            continue

        idx_id = generate_safe_id() 
        doc_payload = {
            "@search.action": "upload",
            "id": idx_id,
            "content": doc.get("text_raw", ""), 
            "filepath": blob_url, 
            "title": document_name,
            "chat_id": chat_id,
            "parent_id": parent_id,
            "timestamp": datetime.now(timezone.utc).isoformat(),
            "content_vector": embedding,
            "page_number": page_number,
            "keywords": doc.get("keywords", []),
            "base64_data": "", 
            "content_markdown": doc.get("text_markdown", ""), 
            "base64_imgs_list": doc.get("base64_imgs_list", []),
            "pdf_language": doc.get("detected_language", "en"), 
            "pdf_name": document_name
        }
        dicts_for_indexing.append(doc_payload)

    # Upload to Azure Search
    if not dicts_for_indexing:
        logging.warning(f"upload_index_docling_tesseract_from_url_async: No documents prepared for indexing for {document_name}. Aborting indexing.")
        return False 

    try:
        async with aiohttp.ClientSession() as session:
            all_responses = await upload_batches(session, url, api_key, dicts_for_indexing)  
            logging.info(f"upload_index_docling_tesseract_from_url_async: Document processed and indexed in {len(all_responses)} batches!")
            return True
    except Exception as e:
        logging.error(f"upload_index_docling_tesseract_from_url_async: Error during indexing: {e}")
        return False
###################################################################################################################################################################
###################################################################################################################################################################
######## FINE ELABORAZIONE EXCEL ##################################################################################################################################
###################################################################################################################################################################
######## FUNZIONI DOCLING + TESSERACT ##############################################################################################################################
###################################################################################################################################################################




####################################################################################################################################################################
############################## Chunker ######################################################################################################
####################################################################################################################################################################

def markdown_to_raw(md_text: str) -> str:
    """
    Converts a Markdown string to its raw text equivalent by first converting it to HTML
    and then stripping all HTML tags.

    Args:
        md_text (str): The input string in Markdown format.

    Returns:
        str: The raw text content extracted from the Markdown, with leading/trailing whitespace removed.
    """
    html = markdown2.markdown(md_text)
    raw = re.sub('<[^<]+?>', '', html)
    return raw.strip()


def split_markdown_and_raw(markdown_text, chunk_size=600, chunk_overlap=200):
    """
    Splits a given Markdown text into smaller chunks, preserving header context
    and converting each chunk to raw text.

    The splitting process involves two main steps:
    1. Splitting the Markdown text by specified header levels (H1 and H2).
    2. Further splitting the resulting Markdown sections into fixed-size chunks
       with a defined overlap using a recursive character splitter.

    For each final chunk, both the Markdown content and its raw text equivalent
    are provided, along with any associated headers.

    Args:
        markdown_text (str): The input text in Markdown format.
        chunk_size (int, optional): The desired maximum size for each text chunk. Defaults to 600.
        chunk_overlap (int, optional): The number of characters to overlap between consecutive chunks. Defaults to 200.

    Returns:
        list: A list of dictionaries, where each dictionary represents a chunk and contains:
            - "markdown" (str): The Markdown content of the chunk.
            - "raw" (str): The raw text content of the chunk (HTML tags stripped).
            - "headers" (dict): A dictionary of headers associated with the chunk.

    Logs:
        - INFO: The total number of chunks created.
        - WARNING: If an empty or whitespace-only chunk is encountered and skipped.
    """
    headers_to_split_on = [
        ("#", "Header 1"),
        ("##", "Header 2"),
#        ("###", "Header 3"),
#        ("####", "Header 4"),
    ]

    # MD splits
    markdown_splitter = MarkdownHeaderTextSplitter(
        headers_to_split_on=headers_to_split_on, strip_headers=False,
    )
    md_header_splits = markdown_splitter.split_text(markdown_text)
    md_header_splits
    # Char-level splits
 


    text_splitter = RecursiveCharacterTextSplitter(
        chunk_size=chunk_size, chunk_overlap=chunk_overlap
    )

    # Split
    splits = text_splitter.split_documents(md_header_splits)
    results = []
    for chunk in splits:
        markdown= chunk.page_content
        if not markdown or markdown.strip() == "":
            logging.warning(f"Skipping empty or whitespace chunk. Metadata: {chunk.metadata}")
            continue
        headers= chunk.metadata
        
        raw = markdown_to_raw(markdown)
        results.append({
            "markdown": markdown,
            "raw": raw,
            "headers": headers,
        })

    logging.info(f"Document split into {len(results)} chunks.")
    return results
def _is_table_line(line: str) -> bool:
    l = line.strip()
    if not l:
        return False
    if l.startswith("|"):
        return True
    return ("|" in l and "---" in l)

def _split_markdown_blocks_preserving_tables(md: str) -> List[str]:
    """
    Divide il markdown in blocchi: tabelle (come blocchi unici) e paragrafi/testo normale.
    """
    lines = md.splitlines()
    blocks: List[str] = []
    cur: List[str] = []
    in_table = False

    def flush():
        nonlocal cur
        if cur:
            blocks.append("\n".join(cur).strip("\n"))
            cur = []

    for ln in lines + [""]:  # sentinella
        if _is_table_line(ln):
            if not in_table and cur:
                flush()
            in_table = True
            cur.append(ln)
            continue
        if in_table and ln.strip() == "":
            # fine tabella
            flush()
            in_table = False
            continue
        if in_table:
            cur.append(ln)
            continue
        # fuori tabella
        if ln.strip() == "":
            flush()
        else:
            cur.append(ln)

    return [b for b in blocks if b.strip()]

def _token_len(s: str, enc) -> int:
    return len(enc.encode(s))

def _split_long_table_block(table_md: str, enc, max_tokens: int) -> List[str]:
    """
    Se un blocco tabella eccede max_tokens, lo spezza per righe replicando header+separator.
    """
    lines = [ln for ln in table_md.splitlines() if ln.strip() != ""]
    if not lines:
        return []

    header = lines[0]
    sep = None
    data_start = 1
    if len(lines) > 1 and ("---" in lines[1] or re.match(r"^\s*\|?\s*:?-+.*-+:?\s*\|?\s*$", lines[1] or "")):
        sep = lines[1]
        data_start = 2

    data_rows = lines[data_start:]
    head_block = header if sep is None else header + "\n" + sep

    if _token_len(head_block, enc) >= max_tokens:
        # Fallback: split per caratteri
        chunked: List[str] = []
        buf = ""
        for ch in table_md:
            if _token_len(buf + ch, enc) >= max_tokens:
                chunked.append(buf)
                buf = ""
            buf += ch
        if buf:
            chunked.append(buf)
        return chunked

    out: List[str] = []
    cur: List[str] = []
    cur_tokens = _token_len(head_block + "\n", enc)

    def push_chunk(rows: List[str]):
        block = head_block + ("\n" if rows else "")
        if rows:
            block += "\n".join(rows)
        out.append(block)

    for row in data_rows:
        row_add = (row + "\n")
        row_toks = _token_len(row_add, enc)
        if cur and cur_tokens + row_toks >= max_tokens:
            push_chunk(cur)
            cur = []
            cur_tokens = _token_len(head_block + "\n", enc)

        if row_toks >= max_tokens:
            # Riga enorme: spezza inline
            slice_buf = ""
            for ch in row_add:
                if _token_len(slice_buf + ch, enc) >= max_tokens:
                    if cur:
                        push_chunk(cur)
                        cur = []
                        cur_tokens = _token_len(head_block + "\n", enc)
                    out.append(slice_buf)
                    slice_buf = ""
                slice_buf += ch
            if slice_buf:
                if cur:
                    push_chunk(cur)
                    cur = []
                    cur_tokens = _token_len(head_block + "\n", enc)
                out.append(slice_buf)
            continue

        cur.append(row)
        cur_tokens += row_toks

    if cur:
        push_chunk(cur)

    return out

def chunk_markdown_excel_safe(
    md: str,
    max_tokens: int = 2000,          # allineato ai parametri XLSX usati in get_chunk_docling_document
    overlap_tokens: int = 100,
    enc_name: str = "o200k_base",
) -> List[str]:
    """
    Chunk del Markdown (da Excel) che preserva le tabelle.
    - Non spezza le tabelle; se una tabella è troppo grande, la spezza per righe replicando l'header.
    - Applica overlap solo tra blocchi non-tabella (o tra gruppi misti).
    """
    enc = tiktoken.get_encoding(enc_name)
    raw_blocks = _split_markdown_blocks_preserving_tables(md)

    # Espandi i blocchi "tabella troppo grande"
    blocks: List[str] = []
    for b in raw_blocks:
        first_line = b.splitlines()[0] if b.splitlines() else ""
        if first_line and _is_table_line(first_line) and _token_len(b, enc) > max_tokens:
            blocks.extend(_split_long_table_block(b, enc, max_tokens))
        else:
            blocks.append(b)

    out: List[str] = []
    buf: List[str] = []
    buf_tok = 0

    def add_with_overlap():
        nonlocal buf, buf_tok
        chunk_text = "\n\n".join(buf)
        out.append(chunk_text)
        if overlap_tokens > 0:
            tail = enc.decode(enc.encode(chunk_text)[-overlap_tokens:])
            buf = [tail] if tail else []
            buf_tok = _token_len("\n\n".join(buf), enc) if buf else 0
        else:
            buf, buf_tok = [], 0

    for b in blocks:
        b_add = b if b.endswith("\n\n") else (b + "\n\n")
        t = _token_len(b_add, enc)
        if buf and buf_tok + t > max_tokens:
            add_with_overlap()
        if t >= max_tokens and not _is_table_line(b.splitlines()[0] if b.splitlines() else ""):
            text = b_add
            piece = ""
            for ch in text:
                if _token_len(piece + ch, enc) >= max_tokens:
                    if buf:
                        add_with_overlap()
                    out.append(piece)
                    piece = ""
                piece += ch
            if piece:
                if buf:
                    add_with_overlap()
                out.append(piece)
            continue

        buf.append(b.rstrip("\n"))
        buf_tok += t
    if buf:
        add_with_overlap()

    return out

####PPTX
def _is_list_line(line: str) -> bool:
    l = line.strip()
    if not l:
        return False
    # bullet o numerate
    return bool(re.match(r"^(\s*[-*•–]\s+|\s*\d{1,3}[.)]\s+)", l))

def _split_markdown_blocks_preserving_tables_and_lists(md: str) -> List[str]:
    """
    Blocchi markdown per PPTX: preserva tabelle come blocchi e raggruppa liste contigue.
    Usa header/righe vuote come separatori fuori da tabelle/liste.
    """
    lines = md.splitlines()
    blocks: List[str] = []
    cur: List[str] = []
    in_table = False
    in_list = False

    def flush():
        nonlocal cur, in_list
        if cur:
            blocks.append("\n".join(cur).strip("\n"))
            cur = []
        in_list = False

    for ln in lines + [""]:  # sentinella
        if _is_table_line(ln):
            if not in_table:
                flush()
                in_table = True
            cur.append(ln)
            continue
        if in_table and ln.strip() == "":
            flush()
            in_table = False
            continue
        if in_table:
            cur.append(ln)
            continue

        # fuori tabella
        if _is_list_line(ln):
            if not in_list:
                flush()
                in_list = True
            cur.append(ln)
            continue
        if in_list and (ln.strip() == "" or ln.lstrip().startswith("#")):
            flush()
            in_list = False
            if ln.strip() == "":
                continue  # consumata come separatore
        # header o separatore slide (---)
        if ln.strip() == "" or ln.strip() == "---" or ln.lstrip().startswith("#"):
            flush()
        else:
            cur.append(ln)

    return [b for b in blocks if b.strip()]

def chunk_markdown_pptx_safe(
    md: str,
    max_tokens: int = 1200,
    overlap_tokens: int = 100,
    enc_name: str = "o200k_base",
) -> List[str]:
    """
    Chunk per PPTX: preserva tabelle e blocchi lista, rispetta header/separatori slide.
    """
    enc = tiktoken.get_encoding(enc_name)
    raw_blocks = _split_markdown_blocks_preserving_tables_and_lists(md)

    # Spezza eventuali tabelle troppo grandi riusando la logica XLSX
    blocks: List[str] = []
    for b in raw_blocks:
        first_line = b.splitlines()[0] if b.splitlines() else ""
        if first_line and _is_table_line(first_line) and _token_len(b, enc) > max_tokens:
            blocks.extend(_split_long_table_block(b, enc, max_tokens))
        else:
            blocks.append(b)

    out: List[str] = []
    buf: List[str] = []
    buf_tok = 0

    def add_with_overlap():
        nonlocal buf, buf_tok
        chunk_text = "\n\n".join(buf)
        out.append(chunk_text)
        if overlap_tokens > 0:
            tail = enc.decode(enc.encode(chunk_text)[-overlap_tokens:])
            buf = [tail] if tail else []
            buf_tok = _token_len("\n\n".join(buf), enc) if buf else 0
        else:
            buf, buf_tok = [], 0

    for b in blocks:
        b_add = b if b.endswith("\n\n") else (b + "\n\n")
        t = _token_len(b_add, enc)
        if buf and buf_tok + t > max_tokens:
            add_with_overlap()
        if t >= max_tokens and not _is_table_line(b.splitlines()[0] if b.splitlines() else ""):
            text = b_add
            piece = ""
            for ch in text:
                if _token_len(piece + ch, enc) >= max_tokens:
                    if buf:
                        add_with_overlap()
                    out.append(piece)
                    piece = ""
                piece += ch
            if piece:
                if buf:
                    add_with_overlap()
                out.append(piece)
            continue

        buf.append(b.rstrip("\n"))
        buf_tok += t
    if buf:
        add_with_overlap()

    return out
###

def generate_safe_id():  
    return str(uuid.uuid4()) 


#################################################################################################################################################################
#################################################################################################################################################################
#################################################################################################################################################################
#################################################################################################################################################################
 
async def get_embedding_async(  
    text: str,  
    global_timeout_time: float,  
    client,  # → AsyncAzureOpenAI  
    model_deployment: str,  
    sem_embedding: asyncio.Semaphore,  
    dimensions: int=1536,  
    timeout: int=120,  
) -> Optional[list]:  
 
    if global_timeout_time:  
        time_left = global_timeout_time - time.time()  
        if time_left <= 0:  
            logging.error("get_embedding_async: Global timeout expired.")  
            return []
    else:  
        time_left = timeout  # fallback  
  
    effective_timeout = min(timeout, time_left)  
    try:
        async with sem_embedding:   
            response = await asyncio.wait_for(  
                client.embeddings.create(  
                    input=[text],  
                    dimensions=dimensions,  
                    model=model_deployment,  
                    timeout=effective_timeout  # timeout interno della chiamata  
                ),  
                timeout=effective_timeout     # timeout totale di attesa per la await  
            )  
            return response.data[0].embedding  
    except asyncio.TimeoutError:  
        logging.error("get_embedding_async: Timeout triggered during embeddings.create.")  
        return []
    except Exception as e:  
        logging.error(f"get_embedding_async: Embedding failed: {e.__class__.__name__}: {e}")  
        return [] 
    except asyncio.CancelledError:  
        logging.warning("get_embedding_async: Task was cancelled")  
        return [] 

  

async def chunk_text_async(text:str, max_tokens:int=8000, overlap:int=100, encoding_name:int="cl100k_base"):  
    """  
    Asynchronously splits a text into chunks of up to max_tokens tokens, with overlap.  
  
    Args:  
        text (str): Input text to split.  
        max_tokens (int): Maximum tokens allowed per chunk.  
        overlap (int): Token overlap between chunks.  
        encoding_name (str): Tokenizer encoding name.  
  
    Returns:  
        list[str]: List of text chunks.  
    """  
    encoding = tiktoken.get_encoding(encoding_name)  
    tokens = encoding.encode(text)  
    chunks = []  
    start = 0  
    while start < len(tokens):  
        end = min(start + max_tokens, len(tokens))  
        chunk = encoding.decode(tokens[start:end])  
        chunks.append(chunk)  
        start += max_tokens - overlap  
    return chunks  
  
async def Matryoshka_embedding_async(  
    text:str,  
    client: AsyncAzureOpenAI,  
    model_deployment:str, 
    global_timeout_time:float,
    sem_embedding: asyncio.Semaphore, 
    dimensions:int=1536,  
    chunk_size:int=8000,  
    overlap:int=100, 
):  
    """  
    Asynchronously computes an embedding for long texts by recursively splitting the text into chunks,  
    processing each chunk (and sub-chunk if still too large), and averaging the embeddings into a final vector.  
  
    Args:  
        text (str): The input text to embed.  
        client (AsyncAzureOpenAI): Async Azure OpenAI client for embeddings.  
        model_deployment (str): The specific deployment to use for the embedding model.  
        dimensions (int): Output dimensionality of the embedding.  
        chunk_size (int): Maximum number of tokens in each chunk.  
        overlap (int): Number of overlapping tokens between chunks.  
  
    Returns:  
        list[float]: The averaged embedding vector for the input text.  
    """  
    time_left = global_timeout_time - time.time()  
    if time_left <= 5:  
        logging.warning("Global timeout less than 5 seconds.") 
        return [] 
    encoding = tiktoken.get_encoding("cl100k_base")  
    if len(encoding.encode(text)) <= chunk_size:  
        return await get_embedding_async(text= text,global_timeout_time= global_timeout_time, client= client,model_deployment= model_deployment,dimensions= dimensions,
                                         sem_embedding= sem_embedding)  
  
    chunks = await chunk_text_async(text, max_tokens=chunk_size, overlap=overlap)  
    
    # Lanciare tutte le embedding in parallelo (gather)  
    async def _embed(chunk):  
        if len(encoding.encode(chunk)) <= chunk_size:  
            return await get_embedding_async(text=chunk, global_timeout_time=global_timeout_time, client= client,model_deployment= model_deployment,dimensions= dimensions,
                                         sem_embedding= sem_embedding)  
        else:  
            return await Matryoshka_embedding_async(text=chunk, client= client,model_deployment= model_deployment, global_timeout_time=global_timeout_time, dimensions= dimensions,
                                         sem_embedding= sem_embedding)  
    time_left = global_timeout_time - time.time()  
    if time_left <= 5:  
        logging.warning("Global timeout less than 5 seconds.") 
        return [] 
  

    try:  
        embeddings = await asyncio.wait_for( 
            asyncio.gather(*[_embed(chunk) for chunk in chunks]),
            timeout= time_left
        ) 
        if not embeddings:  
            return [] 
        return np.mean(embeddings, axis=0).tolist() 

    except asyncio.TimeoutError:  
        logging.warning("Matryoshka_embedding_async: Global timeout reached while waiting for embeddings.")  
        # Potresti decidere di restituire ciò che hai raccolto finora, oppure []  
        return ["Blocking error"] 
    except Exception as e:  
        logging.error(f"Matryoshka_embedding_async: Error during batch embedding: {e}")    
        return ["Blocking error"] 
    except asyncio.CancelledError:  
        logging.warning("Matryoshka_embedding_async: Task was cancelled")
        return ["Blocking error"]


#################################################################################################################################################################
#################################################################################################################################################################
#################################################################################################################################################################
#################################################################################################################################################################


#################################################################################################################################################################
#################################################################################################################################################################
#################################################################################################################################################################
#################################################################################################################################################################


#async def download_blob(blob_url, RecordId,  log_suffix, sem_download: asyncio.Semaphore ):  
#    """
#    Asynchronously downloads the content of a blob from a given URL.
#
#    This function uses an asyncio semaphore to control concurrency of downloads.
#    It performs an asynchronous HTTP GET request and returns the content as bytes.
#    Errors during download are logged, and None is returned.
#
#    Args:
#        blob_url (str): The URL of the blob to download.
#        RecordId (str): The ID of the record associated with this download, used for logging.
#        log_suffix (str): A suffix to append to log messages for identification.
#        semaphore (asyncio.Semaphore, optional): An asyncio semaphore to limit concurrent
#                                                 download operations. Defaults to `sem_download`.
#
#    Returns:
#        Optional[bytes]: The content of the blob as bytes if successful, otherwise None.
#    """
#    try:
#        async with sem_download:   
#            async with aiohttp.ClientSession() as session:  
#                async with session.get(blob_url) as response:  
#                    response.raise_for_status()  
#                    content = await response.read()  
#                    return content  # Oppure BytesIO(content), se serve come file-like  
#    except Exception as e:  
#        logging.warning(f"{log_suffix} Skipping due to blob_url fetch error for RecordId={RecordId}: {e}")  
#        return None  # L'errore viene gestito dal chiamante! 



async def get_pdf_filelike_from_url_async(blob_url: str, RecordId: str, sem_download:asyncio.Semaphore,  log_suffix:str= "") -> Optional[io.BytesIO]:
    """
    Asynchronously downloads a PDF from a given URL into a BytesIO object.

    Args:
        blob_url (str): The URL of the PDF blob to download.

    Returns:
        Optional[io.BytesIO]: A BytesIO object containing the PDF's binary data
            if the download is successful, otherwise None.
    """
    try:
        async with sem_download: 
            async with aiohttp.ClientSession() as session:
                async with session.get(blob_url) as response:
                    response.raise_for_status()  # Solleva un'eccezione per codici di stato HTTP 4xx/5xx
                    pdf_bytes = await response.read()
                    return io.BytesIO(pdf_bytes)
    except aiohttp.ClientError as e:
        #logging.error(f"Error downloading PDF from {blob_url}: {e}")
        logging.error(f"{log_suffix}: Skipping due to blob_url:  {blob_url} fetch error for RecordId={RecordId}: {e}") 
        return None
    except Exception as e:
        logging.error(f"{log_suffix}: An unexpected error occurred while downloading RecordId={RecordId} from {blob_url}: {e}")
        return None




async def get_file_with_retry(blob_url: str, sem_download: asyncio.Semaphore,RecordId: str, global_timeout_time: float, max_retries: int = 3, delay: float = 3, timeout: int=120, log_suffix:str= "") -> Optional[io.BytesIO]:  
    """
    Attempts to download a PDF file from a given blob URL with a retry mechanism.

    This function calls `get_pdf_filelike_from_url_async` (assumed to be defined elsewhere)
    to download the PDF. If the download fails, it retries up to `max_retries` times,
    pausing for `delay` seconds between attempts. It uses a semaphore to control
    the concurrency of download operations.

    Args:
        blob_url (str): The URL of the PDF blob to download.
        sem_download (asyncio.Semaphore): An asyncio semaphore to limit concurrent PDF downloads.
        max_retries (int, optional): The maximum number of times to retry the download. Defaults to 3.
        delay (float, optional): The time in seconds to wait between retries. Defaults to 3.

    Returns:
        Optional[io.BytesIO]: A file-like object (BytesIO) containing the PDF content if the download is successful,
                              otherwise None after all retries are exhausted.

    Logs:
        - WARNING: For each failed attempt, indicating the attempt number, retry delay, and URL.
        - ERROR: If all download attempts fail.
    """
# fallback 

    for attempt in range(1, max_retries+1):  
        if global_timeout_time:  
            time_left = global_timeout_time - time.time()  
            if time_left <= 0:  
                logging.error("Global timeout expired.")  
                return None
        else:  
            time_left = timeout  

        try:
            pdf_io = await asyncio.wait_for( get_pdf_filelike_from_url_async(blob_url=blob_url, sem_download=sem_download, RecordId=RecordId, log_suffix=log_suffix),  
            timeout=time_left  
            )  
            if pdf_io:  
                return pdf_io  
        except asyncio.TimeoutError:  
            logging.error(f"get_file_with_retry: Global timeout reached in get_file_with_retry")
            pdf_io = None  
        except Exception as e:
            logging.error(f"get_file_with_retry: Error downloading PDF (attempt {attempt}): {e}")
        except asyncio.CancelledError:  
            logging.warning("get_file_with_retry: Task was cancelled")  
            return None


        logging.warning(f"get_file_with_retry: Attempt {attempt} to download PDF failed. Retrying in {delay}s... (url: {blob_url})")  
        await asyncio.sleep(delay)  
    logging.error(f"get_file_with_retry: All {max_retries} attempts to download PDF from {blob_url} failed.")  
    return None  








async def process_single_value(value:Dict[str, Any],
                        log_suffix: str,
                        client_openai:AsyncAzureOpenAI,
                        sem:asyncio.Semaphore,
                        sem_download: asyncio.Semaphore,
                        sem_ocr:asyncio.Semaphore,
                        sem_embedding:asyncio.Semaphore,
                        #client_mistral: Mistral,
                        text_analytics_client:TextAnalyticsClient,
                        client_ImageAnalysis: ImageAnalysisClient,
                        global_timeout_time: float, 
                        values: List[Dict[str, Any]],
                        tesseract_config: bool):


    # Controlla se c'è ancora tempo prima del timeout globale
    if time.time() >= global_timeout_time:
        logging.error(f"{log_suffix} Global timeout reached before processing")
        return {}
    
    RecordId = value.get('RecordId')  
    blob_url = value.get('blob_url')  
    document_name = value.get('pdf_name')  
    is_processed = value.get('is_processed')  
    chat_id = value.get('chat_id')  
  
    if not is_processed: 
        if document_name.strip().lower().startswith(("http://", "https://")):
            blob_content= io.BytesIO()
        else:
            blob_content = await get_file_with_retry(blob_url= blob_url,
                                        RecordId= RecordId,
                                        log_suffix= log_suffix,
                                        global_timeout_time= global_timeout_time,
                                        sem_download=sem_download)  
            if blob_content is None:  
                return {  
                    "RecordId": RecordId,  
                    "status": "failed",   
                    "blob_url": blob_url, 
                    "pdf_name": document_name,
                    "is_processed": is_processed,
                    "chat_id": chat_id
                }  
    

        #Il recordId è generato in maniera deterministica dal backend e quindi identifica in maniera univoca il documento, nel futuro si pensa di controllare 
        #se il documento è già stato processato in precedenza, per evitare di ripetere l'operazione
        parent_id = RecordId  
        try:  
            url = (  
                f"https://{AZURE_NAME_SEARCH_SERVICE}.search.windows.net/indexes/{AZURE_SEARCH_INDEX}/docs/search.index?api-version=2024-07-01"  
            )  
            partial_results= await upload_index_docling_tesseract_from_url_async(  
                blob_url=blob_url,
                document_io=blob_content,   
                chat_id=chat_id,  
                parent_id=parent_id,  
                url=url,  
                embedding_function=Matryoshka_embedding_async,  
                client_embedding=client_openai,  
                model_deployment=AZURE_OPENAI_EMBEDDING_DEPLOYMENT,  
                client=client_openai,  
                api_key=AZURE_SEARCH_KEY,  
                document_name=document_name,
                record_id=RecordId,
                sem=sem,
                sem_download=sem_download,
                sem_ocr=sem_ocr,
                sem_embedding=sem_embedding,
               #client_mistral= client_mistral,
               tesseract_config= tesseract_config,
                text_analytics_client= text_analytics_client,
                client_ImageAnalysis= client_ImageAnalysis,   
                global_timeout_time= global_timeout_time,
            )  
            logging.info(f'{log_suffix} Successfully processed PDF.') 


            if not partial_results:  
                logging.warning(f"{log_suffix} No results returned from upload_index_file_from_url_async for RecordId={RecordId}.")  
                return {  
                    "RecordId": RecordId,  
                    "status": "failed",  
                    "error": "No results returned from upload_index_file_from_url_async",  
                    "blob_url": blob_url, 
                    "pdf_name": document_name,
                    "is_processed": is_processed,
                    "chat_id": chat_id  
                }



            await wait_for_record_ids_async(results=values,
                                            record_ids=[RecordId],
                                            log_suffix= log_suffix,
                                            timeout_seconds=215,
                                            poll_interval=5,
                                            global_timeout_time= global_timeout_time 
                                            )

            return {  
                "RecordId": RecordId,  
                "status": "ok",  
                "blob_url": blob_url, 
                "pdf_name": document_name,
                "is_processed": is_processed,
                "chat_id": chat_id  
            }
        except Exception as e:  
            logging.info(f"{log_suffix} Error during indexing: {str(e)}")  
            return {  
                "RecordId": RecordId,  
                "status": "failed",  
                "error": str(e),  
                "blob_url": blob_url, 
                "pdf_name": document_name,
                "is_processed": is_processed,
                "chat_id": chat_id  
            }   
    else:
        logging.info(f"{log_suffix} PDF already processed, skipping: {document_name}")  
        return {  
            "RecordId": RecordId,  
            "status": "ok",  
            "blob_url": blob_url, 
            "pdf_name": document_name,
            "is_processed": is_processed,
            "chat_id": chat_id  
        } 


async def process_values(values: List[Dict[str, Any]],
                        log_suffix: str,
                        tesseract_config: bool,
                        client_openai:AsyncAzureOpenAI,
                        sem:asyncio.Semaphore,
                        sem_download: asyncio.Semaphore,
                        sem_ocr:asyncio.Semaphore,
                        sem_embedding:asyncio.Semaphore,
                        #client_mistral: Mistral,
                        text_analytics_client:TextAnalyticsClient,
                        client_ImageAnalysis: ImageAnalysisClient,
                        global_timeout_time: float):  
                # Controlla se c'è ancora tempo prima del timeout globale

    time_left = global_timeout_time - time.time()  
    if time_left <= 5:  
        logging.warning("process_values:Global timeout less than 5 seconds.") 
        return {"results": []} 


    tasks = [  
        asyncio.create_task(
            process_single_value(value, log_suffix, client_openai,
                                 sem,
                                sem_download,
                                sem_ocr,
                                sem_embedding,
                                #client_mistral,
                                text_analytics_client,
                                client_ImageAnalysis,
                                global_timeout_time,
                                  values,
                                  tesseract_config)  
        )
        for value in values  
    ]    
    try:  
        results = await asyncio.wait_for(  asyncio.gather(*tasks), 
                                         timeout= time_left + 10 ##### The 10-second delta was added to allow child coroutines to close.
                                        )   
    except asyncio.TimeoutError:  
        logging.warning("process_values: Global timeout reached while waiting for asyncio.wait_for(  asyncio.gather(*tasks).")  
        for task in tasks:
            if not task.done():
                task.cancel()  # Cancella task ancora in corso
            if task.done() and task.exception():  # ✅ Controlla se è completata E ha eccezioni
                try:
                    task.exception()  # "Consuma" l'eccezione
                except:
                    pass
        # Potresti decidere di restituire ciò che hai raccolto finora, oppure []  
        return {"results": []} 
    except Exception as e:  
        logging.error(f"process_values:Error during asyncio.wait_for(  asyncio.gather(*tasks) {e}")    
        return {"results": []} 
    except asyncio.CancelledError:  
        logging.warning("process_values: Task was cancelled") 
        for task in tasks:
            if not task.done():
                task.cancel()  # Cancella task ancora in corso
            if task.done() and task.exception():  # ✅ Controlla se è completata E ha eccezioni
                try:
                    task.exception()  # "Consuma" l'eccezione
                except:
                    pass
        return {"results": []}



    return {  
        "results": results
    }  














###################################################################################################################################################################
######## DECISIONE SU CHE ELABORAZIONE FARE #######################################################################################################################
###################################################################################################################################################################     
class ResourceManager:
    """
    Manages all resources (clients and semaphores) used by the application,
    ensuring their proper initialization and cleanup.

    This class is designed to be used as an asynchronous context manager (`async with`)
    to ensure that all resources are properly closed upon exiting the block.

    Attributes:
        clients (dict): A dictionary to store initialized clients (e.g., Mistral, Azure OpenAI).
        semaphores (dict): A dictionary to store asyncio semaphores for concurrency control.
        _cleanup_tasks (list): A list to keep track of pending cleanup tasks.
    """
    
    def __init__(self):
        self.clients = {}
        self.semaphores = {}
        self._cleanup_tasks = []
    
    async def __aenter__(self):
        return self
    
    async def __aexit__(self, exc_type, exc_val, exc_tb):
        """
        Asynchronous context manager exit method.
        Ensures that the cleanup function is called upon exiting the `async with` block.
        """
        await self.cleanup()
    
    async def initialize_clients(self) -> bool:
        """
        Initializes all necessary service clients and semaphores for the application.
        Clients for Mistral, Azure OpenAI, Azure Text Analytics, and Azure Image Analysis are initialized.
        The function handles initialization errors for each client and logs warnings
        or critical errors depending on the resource.

        Returns:
            dict: A dictionary indicating the initialization status of each client (True for success, False for failure).
                  If Azure OpenAI client initialization fails, the process cannot proceed,
                  and the status will reflect this critical error.
        """
        
        all_status = {} 
        # Inizializza semafori
        self.semaphores = {
            'main': asyncio.Semaphore(4),
            'download': asyncio.Semaphore(2),
            'ocr': asyncio.Semaphore(4),
            'embedding': asyncio.Semaphore(3)
        }


#        # Mistral Client
#        try:
#            self.clients['mistral'] = Mistral(
#                api_key=MISTRAL_API_KEY,
#                server_url=MISTRAL_ENDPOINT,
#                retry_config=RetryConfig(
#                    strategy="backoff",
#                    backoff=BackoffStrategy(
#                        initial_interval=500,
#                        max_interval=5000,
#                        exponent=2.0,
#                        max_elapsed_time=20000
#                    ),
#                    retry_connection_errors=True
#                ),
#                timeout_ms=300000
#            )
#            logging.info("Mistral client initialized successfully")
#        except Exception as e:
#            self.clients['mistral'] = None
#            logging.warning(f"Mistral client failed: {e} (fallback available, continuing)")  
#            all_status['mistral'] = False  
#        
        # Azure OpenAI Client
        try:
            self.clients['azure_openai'] = AsyncAzureOpenAI(
                api_version=AZURE_OPENAI_API_VERSION,
                azure_endpoint=AZURE_OPENAI_ENDPOINT,
                api_key=AZURE_OPENAI_API_KEY,
                max_retries=5,
                timeout=30.0  # Timeout esplicito
            )
            all_status['azure_openai'] = True
            logging.info("Azure OpenAI client initialized successfully")
            
        except Exception as e:
            self.clients['azure_openai'] = None  
            logging.critical(f"FAILED TO INITIALIZE OpenAI client, cannot proceed: {e}")  
            all_status['azure_openai'] = False
        
        # Text Analytics Client
        try:
            credential_text_analytics = AzureKeyCredential(API_KEY_COGNITIVE_SERVICE)
            self.clients['text_analytics'] = TextAnalyticsClient(
                endpoint=ENDPOINT_COGNITIVE_SERVICE,
                credential=credential_text_analytics
            )
            all_status['text_analytics'] = True
            logging.info("Text Analytics client initialized successfully")
        except Exception as e:
            self.clients['text_analytics'] = None  
            logging.warning(f"Text Analytics client failed: {e} (fallback with YAKE/langid)")  
            all_status['text_analytics'] = False
        
        # Image Analysis Client
        try:
            credential_image_analysis = AzureKeyCredential(API_KEY_COGNITIVE_SERVICE)
            self.clients['image_analysis'] = ImageAnalysisClient(
                endpoint=ENDPOINT_COGNITIVE_SERVICE,
                credential=credential_image_analysis
            )
            all_status['image_analysis'] = True
            logging.info("Image Analysis client initialized successfully")
        except Exception as e:
            self.clients['image_analysis'] = None  
            logging.warning(f"Image Analysis client failed: {e} (fallback generic description only)")  
            
        
        return all_status
    
    async def cleanup(self):
        """
        Performs a forced cleanup of all managed resources.
        This includes canceling pending cleanup tasks, closing open clients
        (especially Azure OpenAI, which requires explicit asynchronous closure),
        and invoking the garbage collector to free memory.
        """
        logging.info("Starting resource cleanup...")
        
        # Cancella task di cleanup pendenti
        for task in self._cleanup_tasks:
            if not task.done():
                task.cancel()
        
        # Pulisci client Azure OpenAI
        if 'azure_openai' in self.clients and self.clients['azure_openai']:
            try:
                await self.clients['azure_openai'].close()
                logging.info("Azure OpenAI client closed")
            except Exception as e:
                logging.error(f"Error closing Azure OpenAI client: {e}")
        
        # Pulisci altri client se hanno metodi di chiusura
        for client_name, client in self.clients.items():
            if client and hasattr(client, 'close'):
                try:
                    if asyncio.iscoroutinefunction(client.close):
                        await client.close()
                    else:
                        client.close()
                    logging.info(f"{client_name} client closed")
                except Exception as e:
                    logging.error(f"Error closing {client_name} client: {e}")
        
        # Force garbage collection
        gc.collect()
        
        # Small pause to allow connections to close
        await asyncio.sleep(0.1)
        
        logging.info("Resource cleanup completed")


def validate_request_payload(req_body: Dict[str, Any]) -> bool:
    """
    Validates that the request payload has the expected structure for idempotent processing.
    
    Args:
        req_body: The request body to validate
        
    Returns:
        bool: True if valid, False otherwise
    """
    if not isinstance(req_body, dict):
        return False
    
    if 'values' not in req_body:
        return False
    
    values = req_body['values']
    if not isinstance(values, list):
        return False
    
    # Validate each value has required fields
    required_fields = ['RecordId', 'blob_url', 'pdf_name', 'is_processed', 'chat_id']
    for value in values:
        if not isinstance(value, dict):
            return False
        for field in required_fields:
            if field not in value:
                return False
    
    return True











###########################################################################################################################################################################################################
###########################################################################################################################################################################################################
###########################################################################################################################################################################################################
###########################################################################################################################################################################################################


async def process_files_indexing_request(req_body: dict) -> dict:  
    """  
    Funzione batch analoga a main(), ma senza HttpResponse: ritorna solo il JSON dict finale.  
    Args:  
        req_body (dict): payload già decodificato ({ "values": [...] })  
    Returns:  
        dict: { "values": values_processed } come nel body della HttpResponse della main  
    """  
    log_suffix = 'IdemAIProcessFilesUploadIndexFunction.py: '  
    logging.warning("Activity chiamata (no HTTP trigger, pure Python)!")  
  
    global_start_time = time.time()  
    global_timeout_time = global_start_time + 3600.0  
  

    tesseract_config= False
    # Config Tesseract (best-effort)
    try:
        logging.info(f"{log_suffix}: avvio configurazione Tesseract")
        print(f"{log_suffix}: avvio configurazione Tesseract")
        tesseract_config = _configure_tesseract()
        logging.info(f"{log_suffix}: configurazione Tesseract completata")
        print(f"{log_suffix}: configurazione Tesseract completata")
    except Exception as e:
        logging.warning(f"{log_suffix}: Tesseract config warn: {e}")
        print(f"{log_suffix}: Tesseract config warn: {e}")

    try:  
        logging.info(f'{log_suffix} Triggered process_pdf_indexing_request().')  
        values = req_body.get('values', [])  
        # pattern: la funzione process_values modifica 'values' in-place sulle is_processed etc.  
                # Validate request payload structure for idempotency
        if not validate_request_payload(req_body):
            logging.error(f'{log_suffix}:Invalid request payload structure. No Retry')
            return req_body
  
        async with ResourceManager() as resource_manager:  
            try:  
                status_clients = await resource_manager.initialize_clients()  
  
                if not status_clients['azure_openai']:  
                    logging.critical("Critical: Azure OpenAI (embedding/chat) not initialized. Aborting process!")  
                    # Come main(), ritorna l'input come body, status semi-error  
                    return req_body  
  
                # Optional: warning per client non essenziali  
                optional_failed = [key for key, val in status_clients.items() if not val and key != 'azure_openai']  
                if optional_failed:  
                    logging.warning(f"Some optional clients failed to initialize (fallbacks will be used): {optional_failed}")  
  
                # Timeout check  
                if time.time() >= global_timeout_time:  
                    logging.error(f"{log_suffix} Global timeout reached before processing")  
                    return req_body  

                #### devo passare la config di tesseract a tutti 
                try:  
                    results =await process_values(  
                        values=values,  
                        log_suffix=log_suffix,
                        tesseract_config= tesseract_config,  
                        client_openai=resource_manager.clients.get('azure_openai'),  
                        sem=resource_manager.semaphores['main'],  
                        sem_download=resource_manager.semaphores['download'],  
                        sem_ocr=resource_manager.semaphores['ocr'],  
                        sem_embedding=resource_manager.semaphores['embedding'],  
                        #client_mistral=resource_manager.clients.get('mistral'),  
                        text_analytics_client=resource_manager.clients.get('text_analytics'),  
                        client_ImageAnalysis=resource_manager.clients.get('image_analysis'),  
                        global_timeout_time=global_timeout_time  
                    )  
                except Exception as e:  
                    logging.error(f"{log_suffix} Error in process_values: {e}")  
                    # Stessa logica: ritorna il payload originale in caso di errori gravi  
                    return req_body  

                if results["results"]== []: 
                    logging.warning(f"{log_suffix}: No file was elaborated in this request because an exception was raised in the process_values function.")
  
                final_return = {"values": values}  

                elapsed_time = time.time() - global_start_time  
                logging.info(f'{log_suffix} Successfully processed PDF in {elapsed_time:.2f} seconds.')  
                return final_return  
  
            except Exception as inner_e:  
                logging.exception(f"{log_suffix} Error within ResourceManager context: {inner_e}")  
                return req_body  
  
    except Exception as outer_e:  
        logging.exception(f"{log_suffix} Unhandled exception in process_pdf_indexing_request: {outer_e}")  
        return req_body  

